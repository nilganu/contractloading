"""Per-hotel file splitters.

For multi-hotel contracts we want each per-hotel extraction call to see
ONLY that hotel's data — both to avoid LLM laziness and to keep input
token costs proportional to the work done (a 16-hotel Excel becomes 16
single-sheet files, not 16 × the whole workbook).

- Excel (.xlsx/.xls) → one sheet per output file, formatting preserved.
- PDF                 → one page range per output file.
- Anything else       → byte-for-byte copy (LLM call with the whole file).
"""
from __future__ import annotations

import logging
import re
import shutil
from copy import copy
from pathlib import Path
from typing import List, Optional, Tuple

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------


def split_excel_sheet(
    src: Path,
    sheet_name: str,
    dest: Path,
) -> Path:
    """Write a new .xlsx containing ONLY ``sheet_name`` from ``src``.

    Cell values, basic formatting, merged cells and column widths are
    preserved so the LLM sees the same layout as the original.
    """
    import openpyxl

    wb = openpyxl.load_workbook(src)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {src.name}. "
            f"Available: {wb.sheetnames}"
        )
    src_ws = wb[sheet_name]

    new_wb = openpyxl.Workbook()
    # Remove the auto-created blank sheet.
    new_wb.remove(new_wb.active)
    new_ws = new_wb.create_sheet(sheet_name)

    # Copy cells with values + styles.
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(
                row=cell.row, column=cell.column, value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # Merged cells.
    for mr in src_ws.merged_cells.ranges:
        new_ws.merge_cells(str(mr))

    # Column widths.
    for col, dim in src_ws.column_dimensions.items():
        new_ws.column_dimensions[col].width = dim.width

    new_wb.save(dest)
    return dest


def list_excel_sheets(src: Path) -> List[str]:
    """All sheet names in the workbook, in order."""
    import openpyxl

    wb = openpyxl.load_workbook(src, read_only=True)
    return list(wb.sheetnames)


# Sheet names that almost always denote an index / cover / contents sheet,
# not a hotel. Volonline's "Hotel List" is the canonical example.
_INDEX_SHEET_PATTERNS = re.compile(
    r"^\s*("
    r"hotel\s*list|hotels|index|summary|contents|table\s+of\s+contents|"
    r"toc|overview|cover|legend|key|info|info\s*page|notes|terms|"
    r"glossary|master|masterdata|sheet\d+"
    r")\s*$",
    re.IGNORECASE,
)


def is_index_sheet(name: str) -> bool:
    """Heuristic: does this sheet name look like an index/cover, not a hotel?"""
    return bool(_INDEX_SHEET_PATTERNS.match(name or ""))


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------


def list_pdf_page_count(src: Path) -> int:
    from pypdf import PdfReader

    return len(PdfReader(str(src)).pages)


def split_pdf_pages(
    src: Path,
    pages: List[int],  # 1-based page numbers
    dest: Path,
) -> Path:
    """Write a new PDF containing only the given 1-based pages from ``src``."""
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(str(src))
    writer = PdfWriter()
    total = len(reader.pages)
    for p in pages:
        if 1 <= p <= total:
            writer.add_page(reader.pages[p - 1])
    with dest.open("wb") as f:
        writer.write(f)
    return dest


def parse_page_hint(hint: Optional[str], page_count: int) -> List[int]:
    """Best-effort parse of strings like 'Pages 1-3', 'Page 4', '2,5,7'.

    Returns a list of 1-based page numbers. Empty / unparseable hints fall
    back to ALL pages (the caller decides whether to fall through to a
    whole-file call)."""
    if not hint:
        return list(range(1, page_count + 1))
    out: List[int] = []
    # Find ranges like '1-3' and singles like '4'.
    for m in re.finditer(r"(\d+)\s*-\s*(\d+)|(\d+)", hint):
        if m.group(1) and m.group(2):
            lo, hi = int(m.group(1)), int(m.group(2))
            if lo > hi:
                lo, hi = hi, lo
            for p in range(lo, hi + 1):
                if 1 <= p <= page_count and p not in out:
                    out.append(p)
        elif m.group(3):
            p = int(m.group(3))
            if 1 <= p <= page_count and p not in out:
                out.append(p)
    return out or list(range(1, page_count + 1))


# --------------------------------------------------------------------------
# Generic copy (fallback for non-splittable formats)
# --------------------------------------------------------------------------


def copy_file(src: Path, dest: Path) -> Path:
    shutil.copy(str(src), str(dest))
    return dest


# --------------------------------------------------------------------------
# Per-hotel sub-file selection
# --------------------------------------------------------------------------


def split_for_hotel(
    src: Path,
    hotel_name: str,
    source_hint: Optional[str],
    dest_dir: Path,
) -> Tuple[Path, str]:
    """Create a temp sub-file containing just the data for ``hotel_name``.

    Returns ``(sub_file_path, mode)`` where ``mode`` is one of:
    - ``"excel_sheet"`` — single-sheet xlsx
    - ``"pdf_pages"``   — single page-range pdf
    - ``"whole"``       — couldn't split, sent the full file as-is

    Falls back to a whole-file copy if the hint can't be resolved.
    """
    ext = src.suffix.lower()
    dest_dir.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", hotel_name)[:60]

    if ext in (".xlsx", ".xls"):
        # Match the hint to one of the workbook's sheet names.
        try:
            sheets = list_excel_sheets(src)
        except Exception:  # noqa: BLE001
            sheets = []
        sheet = _match_sheet(sheets, hotel_name, source_hint)
        if sheet:
            dest = dest_dir / f"sub-{safe}.xlsx"
            try:
                split_excel_sheet(src, sheet, dest)
                return dest, "excel_sheet"
            except Exception as e:  # noqa: BLE001
                logger.warning(
                    "excel split failed for %r (%s); falling back to whole file",
                    hotel_name, e,
                )

    elif ext == ".pdf":
        try:
            n = list_pdf_page_count(src)
            pages = parse_page_hint(source_hint, n)
            if pages and len(pages) < n:
                dest = dest_dir / f"sub-{safe}.pdf"
                split_pdf_pages(src, pages, dest)
                return dest, "pdf_pages"
        except Exception as e:  # noqa: BLE001
            logger.warning(
                "pdf split failed for %r (%s); falling back to whole file",
                hotel_name, e,
            )

    # Fallback: send the whole file.
    dest = dest_dir / f"sub-{safe}{ext}"
    copy_file(src, dest)
    return dest, "whole"


def _match_sheet(
    sheets: List[str], hotel_name: str, hint: Optional[str]
) -> Optional[str]:
    """Pick the sheet most likely to contain this hotel's data.

    Priority: explicit hint sheet name > exact case-insensitive match >
    fuzzy substring match > None.
    """
    if not sheets:
        return None
    # Hint may look like "Sheet:Barceló Tiran Sharm Resort"
    if hint:
        cleaned = hint.split(":", 1)[-1].strip()
        for s in sheets:
            if s.strip().lower() == cleaned.lower():
                return s
    needle = hotel_name.strip().lower()
    # Exact case-insensitive match
    for s in sheets:
        if s.strip().lower() == needle:
            return s
    # Substring either direction (drop punctuation/whitespace for robustness)
    def _norm(x: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", x.lower())
    n_needle = _norm(hotel_name)
    for s in sheets:
        ns = _norm(s)
        if n_needle and (n_needle in ns or ns in n_needle):
            return s
    return None
