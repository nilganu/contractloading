"""XLSX exporter.

Three modes:
- dynamic_review : exposes UI-internal columns (_source_refs, _confidence, _warnings)
- dynamic_export : dynamic child columns from contract, no internal columns
- strict_template: only the fixed template columns; unsupported child age bands
                   are moved to Extraction Notes with a warning

Always exports two sheets: Hotel and Extraction Notes.

Date cells -> real Excel date values
Numeric cells -> real numbers
Empty values -> blank cells (not "null"/"None"/"")
"""
from __future__ import annotations

import re
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schemas.models import (
    EXTRACTION_NOTES_HEADERS,
    FIXED_BASE_HEADERS,
    FIXED_SUPP_HEADERS,
    STRICT_TEMPLATE_CHILD_COLUMNS,
)
from .moonstride_templates import TEMPLATES as MOONSTRIDE_TEMPLATES
from .moonstride_templates import export_moonstride


_INTERNAL_HEADERS = ["_source_refs", "_confidence", "_warnings"]
# "Days" is NOT numeric — it's a weekday-mask string like "0,1,2,3,4,5,6".
_NUMERIC_HEADERS = {
    "Latitude", "Longitude", "Min Adult", "Max Adult", "Max Pax",
    "Min Stay", "Booking Limit", "Release Period", "Add Charge Value",
    "Charge", "SGL", "DBL", "TPL", "QDP", "Extra Bed",
    "SUPP-HB-ADULT", "SUPP-HB-CHILD", "SUPP-AI-ADULT", "SUPP-AI-CHILD",
}
_DATE_HEADERS = {"Start Date", "End Date"}

_HEADER_FILL = PatternFill(start_color="FFE3F2FD", end_color="FFE3F2FD", fill_type="solid")
_HEADER_FONT = Font(bold=True)

# Excel sheet names: max 31 chars; reserved characters can't be used.
_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sanitize_sheet_name(name: str, *, taken: set[str]) -> str:
    base = _INVALID_SHEET_CHARS.sub("_", (name or "").strip()) or "Hotel"
    base = base[:31]
    candidate = base
    i = 2
    while candidate in taken or candidate.lower() == "hotel" and base.lower() != "hotel":
        suffix = f" ({i})"
        candidate = (base[: 31 - len(suffix)] + suffix)
        i += 1
    taken.add(candidate)
    return candidate


def _date_or_none(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _coerce_for_excel(header: str, value: Any) -> Any:
    if value is None:
        return None
    if header in _DATE_HEADERS:
        return _date_or_none(value)
    if header in _NUMERIC_HEADERS:
        if isinstance(value, (int, float)):
            return value
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _resolve_headers(
    mode: str,
    dynamic_child_columns: List[Dict[str, Any]],
    include_internal: bool,
) -> Tuple[List[str], Dict[str, str]]:
    """Return (header_list, label_to_key_map).

    The header_list contains user-facing labels (so percentage columns show a
    "%" hint). label_to_key_map lets row writers look up the underlying child
    column key from the displayed header.
    """
    label_to_key: Dict[str, str] = {}
    if mode == "strict_template":
        child_headers = list(STRICT_TEMPLATE_CHILD_COLUMNS)
        for h in child_headers:
            label_to_key[h] = h
    else:
        child_headers = []
        for c in dynamic_child_columns:
            label = c.get("label") or c.get("key")
            child_headers.append(label)
            label_to_key[label] = c.get("key") or label

    headers = list(FIXED_BASE_HEADERS) + list(child_headers) + list(FIXED_SUPP_HEADERS)
    if include_internal:
        headers += _INTERNAL_HEADERS
    return headers, label_to_key


def _write_hotel_sheet(
    ws,
    rows: List[Dict[str, Any]],
    headers: List[str],
    label_to_key: Dict[str, str],
    *,
    mode: str,
    source_file: str,
    extra_notes_from_strict: List[Dict[str, Any]],
) -> None:
    """Write the Moonstride header + the supplied rows onto an open sheet."""
    ws.append(headers)
    for col_idx, _header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")

    for row in rows:
        out_row: List[Any] = []
        for header in headers:
            if header in FIXED_BASE_HEADERS or header in FIXED_SUPP_HEADERS:
                out_row.append(_coerce_for_excel(header, row.get(header)))
            elif header.startswith("CHD"):
                dynamic = row.get("dynamicChildValues") or {}
                child_key = label_to_key.get(header, header)
                if mode == "strict_template":
                    if child_key in dynamic:
                        out_row.append(_coerce_for_excel(header, dynamic.get(child_key)))
                    else:
                        out_row.append(None)
                else:
                    out_row.append(_coerce_for_excel(header, dynamic.get(child_key)))
            elif header == "_source_refs":
                out_row.append("; ".join(row.get("_sourceRefs") or []))
            elif header == "_confidence":
                out_row.append(row.get("_confidence"))
            elif header == "_warnings":
                out_row.append("; ".join(row.get("_warnings") or []))
            else:
                out_row.append(None)
        ws.append(out_row)

        if "Start Date" in headers:
            ws.cell(row=ws.max_row, column=headers.index("Start Date") + 1).number_format = "yyyy-mm-dd"
        if "End Date" in headers:
            ws.cell(row=ws.max_row, column=headers.index("End Date") + 1).number_format = "yyyy-mm-dd"

        if mode == "strict_template":
            dynamic = row.get("dynamicChildValues") or {}
            for k, v in dynamic.items():
                if k not in STRICT_TEMPLATE_CHILD_COLUMNS and v is not None:
                    extra_notes_from_strict.append(
                        {
                            "id": f"note_{uuid.uuid4().hex[:8]}",
                            "Source File": source_file,
                            "Page": row.get("sourceSheetOrPage", "—"),
                            "Category": "Child policy",
                            "Note": (
                                f"Strict template export dropped child column '{k}' "
                                f"(value: {v}) for hotel {row.get('Hotel Name')} / "
                                f"{row.get('Room Name')} {row.get('Start Date')} – {row.get('End Date')}"
                            ),
                            "_sourceRefs": row.get("_sourceRefs") or [],
                            "_confidence": row.get("_confidence", 0.5),
                            "hotelName": row.get("Hotel Name"),
                            "linkedHotelRowId": row.get("id"),
                        }
                    )

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    ws.freeze_panes = "A2"


def export_workbook(
    result: Dict[str, Any],
    *,
    output_path: str | Path,
    mode: str = "dynamic_export",
    include_internal: bool = False,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Moonstride deliverable modes write into a copy of the real template.
    # "moonstride_auto" picks the template from the contract's Rate Type.
    if mode == "moonstride_auto" or mode in MOONSTRIDE_TEMPLATES:
        template_id = None if mode == "moonstride_auto" else mode
        export_moonstride(result, output_path=output_path, template_id=template_id)
        return output_path

    if mode == "dynamic_review":
        include_internal = True

    wb = Workbook()
    # First sheet: the combined "Hotel" sheet (every hotel's rows in one
    # place — Moonstride-import friendly).
    combined_ws = wb.active
    combined_ws.title = "Hotel"

    dynamic_child_columns = (result.get("dynamicColumns") or {}).get("childColumns") or []
    headers, label_to_key = _resolve_headers(mode, dynamic_child_columns, include_internal)
    extra_notes_from_strict: List[Dict[str, Any]] = []
    source_file = result.get("workbookSummary", {}).get("sourceFile", "—")
    all_rows = result.get("hotelRows", []) or []

    _write_hotel_sheet(
        combined_ws,
        all_rows,
        headers,
        label_to_key,
        mode=mode,
        source_file=source_file,
        extra_notes_from_strict=extra_notes_from_strict,
    )

    # Per-hotel sheets: one sheet per distinct Hotel Name, preserving input order.
    rows_by_hotel: Dict[str, List[Dict[str, Any]]] = {}
    hotel_order: List[str] = []
    for r in all_rows:
        name = r.get("Hotel Name") or "Unknown Hotel"
        if name not in rows_by_hotel:
            rows_by_hotel[name] = []
            hotel_order.append(name)
        rows_by_hotel[name].append(r)

    if len(hotel_order) > 1:
        taken: set[str] = {"Hotel", "Extraction Notes"}
        for hotel_name in hotel_order:
            sheet_name = _sanitize_sheet_name(hotel_name, taken=taken)
            hotel_ws = wb.create_sheet(sheet_name)

            # Per-hotel: drop CHD columns whose value is null/empty in
            # every row for this hotel. Keeps the per-hotel view tight
            # without affecting the combined Hotel sheet (which still
            # carries the full union of CHD columns for Moonstride).
            this_hotel_rows = rows_by_hotel[hotel_name]
            used_child_keys: set[str] = set()
            for r in this_hotel_rows:
                dyn = r.get("dynamicChildValues") or {}
                for k, v in dyn.items():
                    if v not in (None, ""):
                        used_child_keys.add(k)
            # Strip headers whose backing key isn't in used_child_keys
            filtered_headers: List[str] = []
            for h in headers:
                if not (isinstance(h, str) and h.startswith("CHD")):
                    filtered_headers.append(h)
                    continue
                key = label_to_key.get(h, h)
                if key in used_child_keys:
                    filtered_headers.append(h)
            _write_hotel_sheet(
                hotel_ws,
                this_hotel_rows,
                filtered_headers,
                label_to_key,
                mode=mode,
                source_file=source_file,
                extra_notes_from_strict=extra_notes_from_strict,
            )

    # Extraction Notes sheet
    notes_ws = wb.create_sheet("Extraction Notes")
    notes_ws.append(EXTRACTION_NOTES_HEADERS)
    for col_idx, _ in enumerate(EXTRACTION_NOTES_HEADERS, start=1):
        c = notes_ws.cell(row=1, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT

    all_notes: List[Dict[str, Any]] = list(result.get("extractionNotes") or [])
    all_notes.extend(extra_notes_from_strict)

    for note in all_notes:
        notes_ws.append(
            [
                note.get("Source File") or "",
                str(note.get("Page") or ""),
                note.get("Category") or "",
                note.get("Note") or "",
            ]
        )

    for idx in range(1, len(EXTRACTION_NOTES_HEADERS) + 1):
        notes_ws.column_dimensions[get_column_letter(idx)].width = 30
    notes_ws.column_dimensions[get_column_letter(len(EXTRACTION_NOTES_HEADERS))].width = 80
    notes_ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def export_summary(
    result: Dict[str, Any],
    *,
    mode: str,
    include_internal: bool,
) -> Dict[str, Any]:
    dyn = (result.get("dynamicColumns") or {}).get("childColumns") or []
    headers, _ = _resolve_headers(mode, dyn, include_internal)
    return {
        "headers": headers,
        "hotelRowCount": len(result.get("hotelRows") or []),
        "extractionNotesCount": len(result.get("extractionNotes") or []),
        "dynamicChildColumnCount": len(dyn),
    }
