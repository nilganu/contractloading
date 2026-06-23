"""Moonstride export-template registry + builder.

The final deliverable must match one of three Moonstride hotel-import
templates EXACTLY (same sheet, same header strings, same Rate Type string,
Days as the "1234567" weekday mask). We ship the real template .xlsx files
(``app/templates/moonstride/``) and write rows into a copy of the chosen
template so dropdown validations and the MasterData sheet are preserved.

Template auto-detection is driven by the contract's dominant ``Rate Type``:
  - Per Person ...            -> Per Person Per Night
  - Per Room ... (pax count)  -> Per Room Per Night (Pax count)
  - Per Room ... (otherwise)  -> Per Room Per Night (Adult / Child count)

Internal HotelRow fields map onto template columns:
  SGL/DBL/TPL/QDP -> Adult 1..4 (or 1..4 Pax); dynamic child age-bands ->
  Baby/Child/Teen by age overlap; Extra Bed -> Extra Adult.

Six per-band age columns (Baby/Child/Teen Start/End Age) are appended so a
reviewer can see the contract's actual age bands behind each rate column.
"""
from __future__ import annotations

import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schemas.models import EXTRACTION_NOTES_HEADERS

_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates" / "moonstride"

# Rate-column layouts (in template column order, immediately after "Charge").
_RATE_HEADERS_ADULT_CHILD = [
    "Adult 1 (SGL)",
    "Adult 2 (DBL)",
    "Adult 3 (TPL)",
    "Adult 4 (QUD)",
    "Baby 1 (0-1)",
    "Child 1 (2-12)",
    "Teen 1 (12-17)",
    "Extra Adult",
    "Multi Infant (0-1)",
    "Extra Child (2-12)",
    "Extra Teen (12-17)",
]
_RATE_HEADERS_PAX = [
    "1 Pax",
    "2 Pax",
    "3 Pax",
    "4 Pax",
    "5 Pax",
    "Extra Adult",
    "Multi Infant (0-1)",
    "Extra Child (2-12)",
    "Extra Teen (12-17)",
]

# Position-based child columns appended after the rate columns, mirroring the
# contract layout (Prix lit enfant1 / Age min1 / Age max1 ...): each child
# position gets a Price + Age Min + Age Max triplet. Positions 1-3 are fixed
# so the export schema stays stable; absent positions are left blank.
_CHILD_POSITIONS = [(1, "1st"), (2, "2nd"), (3, "3rd")]
_CHILD_POSITION_HEADERS: List[str] = [
    f"{label} Child {suffix}"
    for _, label in _CHILD_POSITIONS
    for suffix in ("Price", "Age Min", "Age Max")
]

# Map the model's childPosition enum to a 1-based ordinal.
_POSITION_ORDINAL = {"first_child": 1, "second_child": 2, "third_child": 3}

TEMPLATES: Dict[str, Dict[str, Any]] = {
    "moonstride_ppn": {
        "file": "per_person_per_night.xlsx",
        "rate_type": "Per Person Per Night",
        "layout": "adult_child",
    },
    "moonstride_prn_ac": {
        "file": "per_room_adult_child.xlsx",
        "rate_type": "Per Room Per Night (Adult / Child count)",
        "layout": "adult_child",
    },
    "moonstride_prn_pax": {
        "file": "per_room_pax_count.xlsx",
        "rate_type": "Per Room Per Night (Pax count)",
        "layout": "pax",
    },
}

# Moonstride base header -> internal HotelRow key candidates (first hit wins).
# A literal string value means "use this constant default" (handled below).
_BASE_HEADER_TO_KEYS: Dict[str, Tuple[str, ...]] = {
    "Hotel Name": ("Hotel Name",),
    "Hotel Code": ("Hotel Code",),
    "Sell Channel": ("Sell Channel",),
    "Supplier": ("Supplier",),
    "Star Rating": ("Star Rating",),
    "Short Description": ("Short Description",),
    "Bed Type": ("Bed Type",),
    "Max Rollaways": ("Max Rollaways",),
    "Max Cribs (Cots)": ("Max Cribs (Cots)",),
    "Address Line 1": ("Address Line 1",),
    "Address Line 2": ("Address Line 2",),
    "Address Line 3": ("Address Line 3",),
    "Address Line 4": ("Address Line 4",),
    "Postal Code": ("Postal Code",),
    "Country Code": ("Country Code ", "Country Code"),
    "County / State / Province": ("State / Province / Region", "County / State / Province"),
    "City / Area": ("City / Area",),
    "Phone Number": ("Phone Number",),
    "Email Address": ("Email Address",),
    "Hotel Website": ("Hotel Website",),
    "Latitude": ("Latitude",),
    "Longitude": ("Longitude",),
    "Check-In": ("Check-In",),
    "Check-Out": ("Check-Out",),
    "Currency": ("Currency",),
    "Room Name": ("Room Name",),
    "Min Adult": ("Min Adult",),
    "Max Adult": ("Max Adult",),
    "Max Pax": ("Max Pax",),
    "Season": ("Season",),
    "Start Date": ("Start Date",),
    "End Date": ("End Date",),
    "Min Stay": ("Min Stay",),
    "Rate Plan": ("Rate Plan",),
    "Meal Plan": ("Meal Plan",),
    "Status": ("Status",),
    "Booking Limit": ("Booking Limit",),
    "Release Period": ("Release Period",),
    "Customer Price Currency": ("Customer Price Currency",),
    "Add Charge Type": ("Add Charge Type",),
    "Add Charge Value": ("Add Charge Value",),
    "Charge": ("Charge",),
}

# Constant defaults applied when the internal row has no value (mirrors the
# placeholder defaults baked into the template sample rows).
_DEFAULTS: Dict[str, str] = {
    "Check-In": "00:00",
    "Check-Out": "00:00",
    "Status": "Open",
    "Add Charge Type": "Fixed",
    "Charge": "Mark Up",
}

_DATE_HEADERS = {"Start Date", "End Date"}
_NUMERIC_HEADERS = {
    "Latitude", "Longitude", "Min Adult", "Max Adult", "Max Pax",
    "Max Rollaways", "Max Cribs (Cots)",
    "Min Stay", "Booking Limit", "Release Period", "Add Charge Value",
    *_RATE_HEADERS_ADULT_CHILD, *_RATE_HEADERS_PAX, *_CHILD_POSITION_HEADERS,
}

_HEADER_FILL = PatternFill(start_color="FFE3F2FD", end_color="FFE3F2FD", fill_type="solid")
_HEADER_FONT = Font(bold=True)


# --------------------------------------------------------------------------
# Detection
# --------------------------------------------------------------------------
def detect_template(result: Dict[str, Any]) -> str:
    """Pick the Moonstride template from the contract's dominant Rate Type."""
    rows = result.get("hotelRows") or []
    counts: Counter[str] = Counter()
    for r in rows:
        rt = str(r.get("Rate Type") or "").strip().lower()
        if rt:
            counts[rt] += 1
    dominant = counts.most_common(1)[0][0] if counts else ""
    if "per room" in dominant or "per-room" in dominant:
        if "pax" in dominant:
            return "moonstride_prn_pax"
        return "moonstride_prn_ac"
    return "moonstride_ppn"


# --------------------------------------------------------------------------
# Days weekday-mask conversion
# --------------------------------------------------------------------------
def _parse_internal_days(days: Any) -> set[int]:
    """Internal mask uses 0=Sun..6=Sat. Returns the set of present indices."""
    if days is None:
        return set(range(7))
    s = str(days).strip().lower()
    if not s:
        return set(range(7))
    m = re.match(r"^\s*(\d)\s*(?:to|-|–)\s*(\d)\s*$", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return set(range(min(a, b), max(a, b) + 1))
    nums = [int(x) for x in re.findall(r"\d", s) if 0 <= int(x) <= 6]
    return set(nums) if nums else set(range(7))


def days_to_moonstride(days: Any) -> str:
    """Normalize a weekday mask to Moonstride's "1234567" form.

    Moonstride digits are ISO weekdays (1=Mon .. 7=Sun); all-week = "1234567".
    Already-Moonstride values (pure digits 1-7) pass through (deduped/sorted).
    Legacy 0..6 comma masks convert: internal 0=Sun -> 7, 1..6 straight through.
    """
    if days is None:
        return "1234567"
    s = str(days).strip()
    if not s:
        return "1234567"
    if re.fullmatch(r"[1-7]+", s):
        return "".join(sorted(set(s)))
    present = _parse_internal_days(s)
    mapped = sorted({(7 if n == 0 else n) for n in present})
    return "".join(str(n) for n in mapped) or "1234567"


# --------------------------------------------------------------------------
# Child age-band classification
# --------------------------------------------------------------------------
def _classify_band(age_from: Optional[float], age_to: Optional[float]) -> str:
    """Map a contract child age range onto a Moonstride band."""
    if age_to is not None and age_to <= 2:
        return "baby"
    if age_from is not None and age_from >= 12:
        return "teen"
    return "child"


def _band_assignment(
    dynamic_columns: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """Assign dynamic child columns to baby/child/teen bands.

    Within a band the columns keep contract order: the first feeds the
    primary template column, the second feeds the "Multi/Extra" column.
    The band's age range is the min start / max end across its columns.
    """
    bands: Dict[str, Dict[str, Any]] = {
        "baby": {"keys": [], "age_from": None, "age_to": None},
        "child": {"keys": [], "age_from": None, "age_to": None},
        "teen": {"keys": [], "age_from": None, "age_to": None},
    }
    for col in dynamic_columns:
        key = col.get("key")
        if not key:
            continue
        af = col.get("ageFrom")
        at = col.get("ageTo")
        band = _classify_band(af, at)
        slot = bands[band]
        slot["keys"].append(key)
        if af is not None:
            slot["age_from"] = af if slot["age_from"] is None else min(slot["age_from"], af)
        if at is not None:
            slot["age_to"] = at if slot["age_to"] is None else max(slot["age_to"], at)
    for slot in bands.values():
        keys = slot["keys"]
        slot["primary"] = keys[0] if keys else None
        slot["secondary"] = keys[1] if len(keys) > 1 else None
    return bands


def _position_assignment(
    dynamic_columns: List[Dict[str, Any]],
) -> Dict[int, Dict[str, Any]]:
    """Assign dynamic child columns to 1-based positions (1st/2nd/3rd child).

    Honors the column's childPosition when set; otherwise fills the lowest
    free position in contract order. Positions beyond 3 are dropped.
    """
    assigned: Dict[int, Dict[str, Any]] = {}
    leftover: List[Dict[str, Any]] = []
    for col in dynamic_columns:
        if not col.get("key"):
            continue
        pos = _POSITION_ORDINAL.get(col.get("childPosition"))
        if pos and pos not in assigned:
            assigned[pos] = {
                "key": col.get("key"),
                "age_from": col.get("ageFrom"),
                "age_to": col.get("ageTo"),
            }
        else:
            leftover.append(col)
    next_pos = 1
    for col in leftover:
        while next_pos in assigned and next_pos <= 3:
            next_pos += 1
        if next_pos > 3:
            break
        assigned[next_pos] = {
            "key": col.get("key"),
            "age_from": col.get("ageFrom"),
            "age_to": col.get("ageTo"),
        }
        next_pos += 1
    return assigned


# --------------------------------------------------------------------------
# Row building
# --------------------------------------------------------------------------
def _rv(row: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return None


def _date_or_none(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _coerce(header: str, value: Any) -> Any:
    if value is None:
        return None
    if header in _DATE_HEADERS:
        return _date_or_none(value)
    if header in _NUMERIC_HEADERS:
        if isinstance(value, (int, float)):
            return value
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _rate_values_for_row(
    row: Dict[str, Any],
    bands: Dict[str, Dict[str, Any]],
    layout: str,
) -> Dict[str, Any]:
    """Produce the rate-column values for a single row."""
    dyn = row.get("dynamicChildValues") or {}

    def band_val(band: str, slot: str) -> Any:
        key = bands[band].get(slot)
        return dyn.get(key) if key else None

    if layout == "pax":
        return {
            "1 Pax": _rv(row, "SGL"),
            "2 Pax": _rv(row, "DBL"),
            "3 Pax": _rv(row, "TPL"),
            "4 Pax": _rv(row, "QDP"),
            "5 Pax": None,
            "Extra Adult": _rv(row, "Extra Bed"),
            "Multi Infant (0-1)": band_val("baby", "primary"),
            "Extra Child (2-12)": band_val("child", "primary"),
            "Extra Teen (12-17)": band_val("teen", "primary"),
        }
    return {
        "Adult 1 (SGL)": _rv(row, "SGL"),
        "Adult 2 (DBL)": _rv(row, "DBL"),
        "Adult 3 (TPL)": _rv(row, "TPL"),
        "Adult 4 (QUD)": _rv(row, "QDP"),
        "Baby 1 (0-1)": band_val("baby", "primary"),
        "Child 1 (2-12)": band_val("child", "primary"),
        "Teen 1 (12-17)": band_val("teen", "primary"),
        "Extra Adult": _rv(row, "Extra Bed"),
        "Multi Infant (0-1)": band_val("baby", "secondary"),
        "Extra Child (2-12)": band_val("child", "secondary"),
        "Extra Teen (12-17)": band_val("teen", "secondary"),
    }


def build_row_values(
    row: Dict[str, Any],
    bands: Dict[str, Dict[str, Any]],
    positions: Dict[int, Dict[str, Any]],
    *,
    rate_type: str,
    layout: str,
) -> Dict[str, Any]:
    """Map one internal HotelRow to a {header: coerced value} dict."""
    out: Dict[str, Any] = {}

    # Base columns
    for header, keys in _BASE_HEADER_TO_KEYS.items():
        out[header] = _coerce(header, _rv(row, *keys))

    # Overrides / derived base columns. The template's single "Child Age" /
    # "Infant Age" columns are intentionally left blank — the per-position
    # child columns (appended below) are the source of truth.
    out["Rate Type"] = rate_type
    out["Days"] = days_to_moonstride(row.get("Days"))

    # Constant defaults where empty
    for header, default in _DEFAULTS.items():
        if out.get(header) in (None, ""):
            out[header] = default
    if out.get("Currency") in (None, ""):
        out["Currency"] = "EUR"
    if out.get("Customer Price Currency") in (None, ""):
        out["Customer Price Currency"] = out.get("Currency") or "EUR"

    # Template rate columns (filled by age band so the native Moonstride
    # columns import correctly).
    rate_vals = _rate_values_for_row(row, bands, layout)
    for header, value in rate_vals.items():
        out[header] = _coerce(header, value)

    # Appended per-position child columns: Price (this row), Age Min / Age Max
    # (column metadata, shared by every row), mirroring the contract layout.
    dyn = row.get("dynamicChildValues") or {}
    for pos, label in _CHILD_POSITIONS:
        info = positions.get(pos)
        price_h = f"{label} Child Price"
        min_h = f"{label} Child Age Min"
        max_h = f"{label} Child Age Max"
        if info:
            out[price_h] = _coerce(price_h, dyn.get(info["key"]))
            out[min_h] = _coerce(min_h, info["age_from"])
            out[max_h] = _coerce(max_h, info["age_to"])
        else:
            out[price_h] = None
            out[min_h] = None
            out[max_h] = None

    return out


# --------------------------------------------------------------------------
# Workbook export
# --------------------------------------------------------------------------
def _template_path(template_id: str) -> Path:
    spec = TEMPLATES[template_id]
    return _TEMPLATE_DIR / spec["file"]


def _read_header_row(ws) -> List[Optional[str]]:
    return [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]


def export_moonstride(
    result: Dict[str, Any],
    *,
    output_path: str | Path,
    template_id: Optional[str] = None,
) -> Tuple[Path, str]:
    """Write the result into a copy of the chosen Moonstride template.

    Returns (output_path, template_id_used).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if template_id is None or template_id not in TEMPLATES:
        template_id = detect_template(result)
    spec = TEMPLATES[template_id]

    wb = load_workbook(_template_path(template_id))
    ws = wb["Hotel"] if "Hotel" in wb.sheetnames else wb.active

    # Locate template header columns by name.
    header_row = _read_header_row(ws)
    header_to_col: Dict[str, int] = {}
    for idx, h in enumerate(header_row, start=1):
        if isinstance(h, str) and h.strip():
            header_to_col[h.strip()] = idx

    # Append the per-position child headers in the first blank columns after
    # the last populated template column.
    last_col = max((c for h, c in header_to_col.items()), default=ws.max_column)
    next_col = last_col + 1
    for child_header in _CHILD_POSITION_HEADERS:
        if child_header not in header_to_col:
            header_to_col[child_header] = next_col
            cell = ws.cell(row=1, column=next_col, value=child_header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            next_col += 1

    # Clear any template sample data rows below the header.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    dynamic_columns = (result.get("dynamicColumns") or {}).get("childColumns") or []
    bands = _band_assignment(dynamic_columns)
    positions = _position_assignment(dynamic_columns)
    rows = result.get("hotelRows") or []

    write_row = 2
    for row in rows:
        values = build_row_values(
            row, bands, positions, rate_type=spec["rate_type"], layout=spec["layout"]
        )
        for header, value in values.items():
            col = header_to_col.get(header)
            if col is None:
                continue
            cell = ws.cell(row=write_row, column=col, value=value)
            if header in _DATE_HEADERS and value is not None:
                cell.number_format = "yyyy-mm-dd"
        write_row += 1

    _append_notes_sheet(wb, result)
    wb.save(output_path)
    return output_path, template_id


def _append_notes_sheet(wb, result: Dict[str, Any]) -> None:
    if "Extraction Notes" in wb.sheetnames:
        del wb["Extraction Notes"]
    ws = wb.create_sheet("Extraction Notes")
    ws.append(EXTRACTION_NOTES_HEADERS)
    for col_idx, _ in enumerate(EXTRACTION_NOTES_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    for note in result.get("extractionNotes") or []:
        ws.append(
            [
                note.get("Source File") or "",
                str(note.get("Page") or ""),
                note.get("Category") or "",
                note.get("Note") or "",
            ]
        )
    for idx in range(1, len(EXTRACTION_NOTES_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 30
    ws.column_dimensions[get_column_letter(len(EXTRACTION_NOTES_HEADERS))].width = 80
    ws.freeze_panes = "A2"


def write_raw_rows(
    rows: List[Dict[str, Any]],
    template_id: str,
    output_path: str | Path,
) -> Path:
    """Pass-through writer: dump rows GPT already keyed by Moonstride header
    names straight into the chosen template, no field mapping, no
    normalization. Unknown headers in a row are silently dropped.
    """
    if template_id not in TEMPLATES:
        template_id = "moonstride_ppn"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(_template_path(template_id))
    ws = wb["Hotel"] if "Hotel" in wb.sheetnames else wb.active

    # Locate header columns (template header strings include the trailing
    # blanks past the populated set — strip and skip blank/None entries).
    header_row = _read_header_row(ws)
    header_to_col: Dict[str, int] = {}
    for idx, h in enumerate(header_row, start=1):
        if isinstance(h, str) and h.strip():
            header_to_col[h.strip()] = idx

    # Per Jun 2026 user rule: child-band / per-position child columns
    # are NO LONGER produced. Child policy lives in the supplement file
    # only. Skip the auto-append that previously created these columns.

    # Wipe the sample rows the template ships with.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    date_headers = {"Start Date", "End Date"}
    # Track (row, col) of cells the mapper flagged as AI-derived so we
    # can recolour them after the standard colour-strip below.
    ai_cells_to_recolour: List[Tuple[int, int]] = []
    for r_idx, row in enumerate(rows, start=2):
        if not isinstance(row, dict):
            continue
        ai_fields = row.get("_ai_fields") or set()
        for key, value in row.items():
            key_str = str(key).strip()
            if key_str.startswith("_"):
                continue  # side-channel, never written to a cell
            col = header_to_col.get(key_str)
            if col is None:
                continue
            cell = ws.cell(row=r_idx, column=col, value=value)
            if key_str in date_headers and isinstance(value, str) and value:
                cell.number_format = "yyyy-mm-dd"
            if key_str in ai_fields and value not in (None, ""):
                ai_cells_to_recolour.append((r_idx, col))

    # Per Jun 2026 user rule: child-band columns are removed entirely
    # from the hotel rate file (not just left blank). Delete them so
    # downstream Moonstride importer doesn't see them as expected
    # columns. Delete right-to-left so column indices stay stable.
    unwanted_headers = {
        # Per-position child price/age columns (the writer used to
        # auto-append these for the legacy mapper output).
        "1st Child Price", "1st Child Age Min", "1st Child Age Max",
        "2nd Child Price", "2nd Child Age Min", "2nd Child Age Max",
        "3rd Child Price", "3rd Child Age Min", "3rd Child Age Max",
        # Band-classified rate columns shipped in the template.
        "Baby 1 (0-1)", "Child 1 (2-12)", "Teen 1 (12-17)",
        "Multi Infant (0-1)", "Extra Child (2-12)", "Extra Teen (12-17)",
    }
    to_delete: List[int] = []
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if isinstance(header_val, str) and header_val.strip() in unwanted_headers:
            to_delete.append(col_idx)
    for col_idx in reversed(to_delete):
        ws.delete_cols(col_idx)

    # Strip inherited header colours (grey fill + red bold text) from the
    # bundled Moonstride templates — the user wants plain Excel output.
    _no_fill = PatternFill(fill_type=None)
    _plain_header = Font(name="Calibri", size=11, bold=True, color="FF000000")
    _plain_body = Font(name="Calibri", size=11, color="FF000000")
    for cell in ws[1]:
        cell.fill = _no_fill
        cell.font = _plain_header
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.fill = _no_fill
            cell.font = _plain_body

    # Re-paint AI-derived cells AFTER the global colour-strip so the
    # marker survives. Italic medium-blue is visible against the plain
    # black body but doesn't scream like red.
    _ai_font = Font(
        name="Calibri", size=11, italic=True, color="FF0070C0",
    )
    for r_idx, col_idx in ai_cells_to_recolour:
        ws.cell(row=r_idx, column=col_idx).font = _ai_font

    wb.save(output_path)
    return output_path


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    return value


def preview_moonstride(
    result: Dict[str, Any], template_id: Optional[str] = None
) -> Dict[str, Any]:
    """Build the exact export table as JSON (headers + per-row header->value).

    Lets the UI show precisely what the generated Excel will contain, using
    the same mapping as ``export_moonstride`` — without writing a file.
    """
    if template_id is None or template_id not in TEMPLATES:
        template_id = detect_template(result)
    spec = TEMPLATES[template_id]

    wb = load_workbook(_template_path(template_id), read_only=True)
    ws = wb["Hotel"] if "Hotel" in wb.sheetnames else wb.active
    tmpl_headers = [c for c in next(ws.iter_rows(values_only=True)) if c is not None]
    wb.close()
    headers = list(tmpl_headers) + list(_CHILD_POSITION_HEADERS)

    dynamic_columns = (result.get("dynamicColumns") or {}).get("childColumns") or []
    bands = _band_assignment(dynamic_columns)
    positions = _position_assignment(dynamic_columns)

    rows_out: List[Dict[str, Any]] = []
    for row in result.get("hotelRows") or []:
        vals = build_row_values(
            row, bands, positions, rate_type=spec["rate_type"], layout=spec["layout"]
        )
        rows_out.append({h: _json_safe(vals.get(h)) for h in headers})

    return {
        "templateId": template_id,
        "rateType": spec["rate_type"],
        "headers": headers,
        "rows": rows_out,
    }


def moonstride_summary(result: Dict[str, Any], template_id: Optional[str]) -> Dict[str, Any]:
    if template_id is None or template_id not in TEMPLATES:
        template_id = detect_template(result)
    spec = TEMPLATES[template_id]
    layout_rates = (
        _RATE_HEADERS_PAX if spec["layout"] == "pax" else _RATE_HEADERS_ADULT_CHILD
    )
    return {
        "templateId": template_id,
        "rateType": spec["rate_type"],
        "rateColumns": layout_rates,
        "childPositionColumns": list(_CHILD_POSITION_HEADERS),
        "hotelRowCount": len(result.get("hotelRows") or []),
        "extractionNotesCount": len(result.get("extractionNotes") or []),
    }
