"""Excel parser.

Goal: feed downstream stages a rich, faithful, sheet-by-sheet view of the
workbook. We do not try to interpret rate blocks here — that belongs in
ir_builder + the LLM stage.

For every sheet we capture:
- worksheet name
- used range (e.g. A1:S35)
- merged ranges
- visible + hidden row/column flags
- a 2D grid of cells with type, raw value, displayed value, formula,
  and style hints (bold, fill color, has border)
- cell comments
"""
from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..file_type import FileKind


def _cell_value_for_display(value: Any) -> Any:
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return value


def _is_truthy_color(c: Any) -> bool:
    if c is None:
        return False
    rgb = getattr(c, "rgb", None)
    if not rgb or not isinstance(rgb, str):
        return False
    if rgb in {"00000000", "FFFFFFFF"}:
        return False
    return True


def _parse_with_openpyxl(path: Path, source_file: str) -> Dict[str, Any]:
    # data_only=False so we still see formulas, but we want displayed values too.
    # Open twice: once with formulas, once with cached results.
    wb_formulas = load_workbook(path, data_only=False, read_only=False)
    wb_values = load_workbook(path, data_only=True, read_only=False)

    sheets: List[Dict[str, Any]] = []

    for ws_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[ws_name]
        ws_v = wb_values[ws_name]

        min_row, max_row = ws_f.min_row or 1, ws_f.max_row or 1
        min_col, max_col = ws_f.min_column or 1, ws_f.max_column or 1
        used_range = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

        merged = [str(r) for r in ws_f.merged_cells.ranges]

        hidden_rows = [
            r
            for r in range(min_row, max_row + 1)
            if ws_f.row_dimensions.get(r) and ws_f.row_dimensions[r].hidden
        ]
        hidden_cols = [
            get_column_letter(c)
            for c in range(min_col, max_col + 1)
            if ws_f.column_dimensions.get(get_column_letter(c))
            and ws_f.column_dimensions[get_column_letter(c)].hidden
        ]

        rows: List[List[Dict[str, Any]]] = []
        for r in range(min_row, max_row + 1):
            row: List[Dict[str, Any]] = []
            for c in range(min_col, max_col + 1):
                cell_f = ws_f.cell(row=r, column=c)
                cell_v = ws_v.cell(row=r, column=c)
                coord = f"{get_column_letter(c)}{r}"

                raw = cell_f.value
                displayed = _cell_value_for_display(cell_v.value)
                is_formula = isinstance(raw, str) and raw.startswith("=")

                style_hints: List[str] = []
                try:
                    if cell_f.font and cell_f.font.bold:
                        style_hints.append("bold")
                    if cell_f.fill and _is_truthy_color(cell_f.fill.start_color):
                        style_hints.append("filled")
                    if cell_f.border and any(
                        getattr(cell_f.border, side).style
                        for side in ("left", "right", "top", "bottom")
                    ):
                        style_hints.append("bordered")
                except Exception:  # noqa: BLE001 - openpyxl style edge cases
                    pass

                comment = None
                try:
                    if cell_f.comment is not None:
                        comment = cell_f.comment.text
                except Exception:  # noqa: BLE001
                    comment = None

                row.append(
                    {
                        "coord": coord,
                        "row": r,
                        "col": c,
                        "value": displayed,
                        "raw": raw if not is_formula else None,
                        "formula": raw if is_formula else None,
                        "type": type(displayed).__name__ if displayed is not None else None,
                        "style": style_hints or None,
                        "comment": comment,
                    }
                )
            rows.append(row)

        sheets.append(
            {
                "name": ws_name,
                "sheet_state": ws_f.sheet_state,
                "used_range": used_range,
                "merged_ranges": merged,
                "hidden_rows": hidden_rows,
                "hidden_columns": hidden_cols,
                "rows": rows,
            }
        )

    return {
        "source_file": source_file,
        "input_format": "xlsx",
        "sheet_names": wb_formulas.sheetnames,
        "sheets": sheets,
    }


def _parse_with_xlrd(path: Path, source_file: str) -> Dict[str, Any]:
    import xlrd  # type: ignore

    book = xlrd.open_workbook(str(path), formatting_info=False)
    sheets: List[Dict[str, Any]] = []
    for ws in book.sheets():
        if ws.nrows == 0:
            sheets.append(
                {
                    "name": ws.name,
                    "sheet_state": "visible",
                    "used_range": "A1:A1",
                    "merged_ranges": [],
                    "hidden_rows": [],
                    "hidden_columns": [],
                    "rows": [],
                }
            )
            continue
        rows: List[List[Dict[str, Any]]] = []
        for r in range(ws.nrows):
            row: List[Dict[str, Any]] = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(cell.value, book.datemode).isoformat()
                    except Exception:  # noqa: BLE001
                        value = cell.value
                row.append(
                    {
                        "coord": f"{get_column_letter(c + 1)}{r + 1}",
                        "row": r + 1,
                        "col": c + 1,
                        "value": value,
                        "raw": cell.value,
                        "formula": None,
                        "type": type(value).__name__ if value is not None else None,
                        "style": None,
                        "comment": None,
                    }
                )
            rows.append(row)
        used_range = (
            f"A1:{get_column_letter(max(ws.ncols, 1))}{max(ws.nrows, 1)}"
        )
        sheets.append(
            {
                "name": ws.name,
                "sheet_state": "visible",
                "used_range": used_range,
                "merged_ranges": [],
                "hidden_rows": [],
                "hidden_columns": [],
                "rows": rows,
            }
        )

    return {
        "source_file": source_file,
        "input_format": "xls",
        "sheet_names": [s.name for s in book.sheets()],
        "sheets": sheets,
    }


def parse_excel(path: str | Path, *, kind: FileKind = "xlsx") -> Dict[str, Any]:
    p = Path(path)
    source_file = p.name
    if kind == "xls":
        try:
            return _parse_with_xlrd(p, source_file)
        except Exception:  # noqa: BLE001
            # Some .xls files are actually xlsx renamed — try openpyxl.
            return _parse_with_openpyxl(p, source_file)
    return _parse_with_openpyxl(p, source_file)


def cell_at(sheet: Dict[str, Any], coord: str) -> Optional[Dict[str, Any]]:
    """Look up a cell by A1 coord inside a parsed sheet dict."""
    for row in sheet["rows"]:
        for cell in row:
            if cell["coord"] == coord:
                return cell
    return None


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _fmt_cell_for_grid(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\t", " ").replace("\r\n", " / ").replace("\n", " / ").strip()
    return s


def sheet_text_preview(sheet: Dict[str, Any], max_rows: int | None = None) -> str:
    """Render a sheet as a column-letter-tagged grid the LLM can read like Excel.

    Format:
        === Sheet: <name> ===
        Row    A           B           C       ...
        1      Hotel       …           …
        2      Currency    EUR
        ...

    - Each row is prefixed by its 1-based row number so the LLM can
      reference specific cells back in extraction notes.
    - Column headers are Excel column letters (A, B, …, Z, AA, …).
    - Trailing empty columns and rows are trimmed.
    - Fully-blank intermediate rows are skipped to save tokens.
    """
    name = sheet.get("name") or "Sheet"
    rows = sheet.get("rows", [])
    if max_rows is not None:
        rows = rows[:max_rows]

    if not rows:
        return f"=== Sheet: {name} ===\n(empty)"

    # Build a 2D string grid first so we can trim accurately.
    grid: List[List[str]] = []
    for r in rows:
        grid.append([_fmt_cell_for_grid(c.get("value")) for c in r])

    # Trim trailing columns
    max_used_col = 0
    for row in grid:
        for i in range(len(row) - 1, -1, -1):
            if row[i]:
                max_used_col = max(max_used_col, i + 1)
                break
    if max_used_col == 0:
        return f"=== Sheet: {name} ===\n(empty)"

    # Find last used row
    last_used_row = 0
    for r, row in enumerate(grid):
        if any(c for c in row[:max_used_col]):
            last_used_row = r + 1

    cols = [_col_letter(i + 1) for i in range(max_used_col)]
    lines = [f"=== Sheet: {name} ==="]
    lines.append("Row\t" + "\t".join(cols))
    for r in range(last_used_row):
        cells = grid[r][:max_used_col] if r < len(grid) else [""] * max_used_col
        if not any(c for c in cells):
            continue  # skip blank rows inside the sheet
        lines.append(f"{r + 1}\t" + "\t".join(cells))
    return "\n".join(lines)
