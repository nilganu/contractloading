"""Supplement-template writer.

Loads ``backend/app/templates/supplement/hotel-supplement.xlsx`` and
writes one row per dict supplied. The mapper has already applied every
conditional rule and forced value, so this writer is intentionally
minimal — just header lookup + cell writes.

NOTE on the bundled template: use ``openpyxl.load_workbook(..., read_only=False)``
(the default). Read-only mode silently truncates ``max_column`` to the
last consistently populated column across the iter buffer and HIDES
the three leading columns (``Hotel Name`` / ``Hotel Code`` /
``Suppliment Code``) on this header-only template — which is exactly
how the first inspection of the file missed them.

The header names use the literal typo ``Suppliment`` (not "Supplement").
Do not "fix" the typo — the Moonstride importer matches by exact string.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from typing import Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Reused to wipe inherited header colours: no fill, default bold black.
_NO_FILL = PatternFill(fill_type=None)
_PLAIN_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FF000000")
_PLAIN_BODY_FONT = Font(name="Calibri", size=11, color="FF000000")


def _strip_colours(ws) -> None:
    """Remove fill + font colour from every used cell so the output is a
    plain Excel file (Moonstride doesn't need colours and the user
    explicitly asked for them off)."""
    if ws.max_row == 0:
        return
    for cell in ws[1]:
        cell.fill = _NO_FILL
        cell.font = _PLAIN_HEADER_FONT
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = _NO_FILL
            cell.font = _PLAIN_BODY_FONT

_TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "templates"
    / "supplement"
    / "hotel-supplement.xlsx"
)

# Exact header strings, in template order. Matches the bundled
# Hotel-Supplement-Import-Sample-AI.xlsx (Sheet name: 'Supplement').
SUPPLEMENT_HEADERS: List[str] = [
    "Hotel Name",
    "Hotel Code",
    "Supplement Code",
    "Supplement Name",
    "Rooms",
    "Rate Plans",
    "Min Stay",
    "Max Stay",
    "Supplement Type",
    "Display As Separate Room",
    "Meal Plan",
    "Required Supplement",
    "Restricted Supplement",
    "Display on Customer Documentation",
    "Display on Supplier Notification",
    "Description",
    "Contract Period",
    "Season Name",
    "Start Date (DD-MM-YYYY)",
    "End Date (DD-MM-YYYY)",
    "Supplier",
    "Currency",
    "Charge Type",
    "Calculation Method",
    "Traveler Type",
    "FareType Name",
    "Standard / Count / Index",
    "Min Age",
    "Max Age",
    "Min Adult",
    "Max Adult",
    "Max Child",
    "Supplier Cost",
    "Customer Price",
]

_HEADER_FILL = PatternFill(
    start_color="FFE3F2FD", end_color="FFE3F2FD", fill_type="solid"
)
_HEADER_FONT = Font(bold=True)


def _template_path() -> Path:
    return _TEMPLATE_PATH


def write_supplement_rows(
    rows: List[Dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """Write supplement rows into a copy of the bundled template.

    Rows must be keyed by the literal header strings from
    ``SUPPLEMENT_HEADERS``. Unknown keys are silently dropped (mirrors
    ``moonstride_templates.write_raw_rows``). Missing keys land as blank.
    An empty ``rows`` list produces a header-only file — intentional, so
    "no supplements" is visible as an artifact rather than a missing
    download.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(_template_path())
    # The AI-format template uses the sheet name 'Supplement'; older
    # variants used 'Sheet1'. Fall through to active as a last resort.
    if "Supplement" in wb.sheetnames:
        ws = wb["Supplement"]
    elif "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
    else:
        ws = wb.active

    # Build header -> column map. Use default load (NOT read_only) so the
    # three leading columns are visible — see module docstring.
    header_to_col: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip():
            header_to_col[v.strip()] = col

    # Auto-append any headers declared in SUPPLEMENT_HEADERS that the
    # bundled template predates (e.g. Min Adult / Max Adult / Max Child
    # added in Jun 2026). Append at the first blank column after the
    # last populated one, preserving template order for the existing
    # headers.
    last_col = max(header_to_col.values(), default=ws.max_column)
    next_col = last_col + 1
    for header in SUPPLEMENT_HEADERS:
        if header not in header_to_col:
            header_to_col[header] = next_col
            cell = ws.cell(row=1, column=next_col, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            next_col += 1

    # Wipe any sample rows below the header.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Track AI-derived cells so we can re-paint them in italic blue
    # AFTER _strip_colours sets every cell to plain black.
    ai_cells_to_recolour: List[Tuple[int, int]] = []
    for r_idx, row in enumerate(rows, start=2):
        if not isinstance(row, dict):
            continue
        ai_fields = row.get("_ai_fields") or set()
        for key, value in row.items():
            key_str = str(key).strip()
            if key_str.startswith("_"):
                continue  # side-channel, never written to a cell
            col = header_to_col.get(key_str)
            if col is None:
                continue
            ws.cell(row=r_idx, column=col, value=value)
            if key_str in ai_fields and value not in (None, ""):
                ai_cells_to_recolour.append((r_idx, col))

    # Defensive: column-width hint so the file opens reasonably in Excel.
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width in (None, 0):
            ws.column_dimensions[letter].width = 18

    # Strip inherited header colours (grey fill, red bold text).
    _strip_colours(ws)

    # Re-paint AI-derived cells after the strip. Italic medium-blue,
    # mirroring the hotel writer for visual consistency.
    _ai_font = Font(name="Calibri", size=11, italic=True, color="FF0070C0")
    for r_idx, col_idx in ai_cells_to_recolour:
        ws.cell(row=r_idx, column=col_idx).font = _ai_font

    wb.save(output_path)
    return output_path
