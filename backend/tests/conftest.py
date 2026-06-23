"""Pytest fixtures.

Builds a synthetic multi-sheet Excel contract that mirrors the structure of
the Volonline workbook (Hotel List index + per-hotel rate sheets).
"""
from __future__ import annotations

import os
import sys
from pathlib import Path
from typing import Iterator

import pytest
from openpyxl import Workbook

# Ensure backend/ is on sys.path even when running pytest from repo root
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault("DATABASE_URL", "sqlite:///./test_hotel.db")
os.environ.setdefault("STORAGE_DIR", str(ROOT / "storage"))
# Force stub mode in tests — overrides any .env value so we never hit the real
# OpenAI API from automated test runs.
os.environ["OPENAI_API_KEY"] = ""


HOTEL_NAMES = [
    "Barceló Tiran Sharm Resort",
    "Charmillion Club Aqua Park",
    "Mövenpick Resort Sharm El Sheik",
]


def _build_hotel_sheet(wb: Workbook, hotel: str) -> None:
    ws = wb.create_sheet(title=hotel[:31])
    ws.append([hotel])
    ws.append([])
    ws.append(["Currency", "EUR"])
    ws.append(["Meal Plan", "BB"])
    ws.append([])
    ws.append(["FROM", "TO", "Room", "DBL", "SGL", "CHD(2-11.99)", "CHD(0-1.99)"])
    ws.append(["01/05/2025", "31/10/2025", "Standard Double", 120, 180, 60, 0])
    ws.append(["01/11/2025", "31/03/2026", "Standard Double", 90, 130, 45, 0])
    ws.append([])
    ws.append(["Cancellation: 30 days prior to arrival no charge, 15 days 50%"])
    ws.append(["Gala dinner 24/12: 80 EUR per adult, 40 EUR per child"])


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    # Hotel List index
    index = wb.create_sheet("Hotel List")
    index.append(["Hotel", "Destination", "Country"])
    for h in HOTEL_NAMES:
        index.append([h, "Sharm El Sheikh", "EG"])

    for h in HOTEL_NAMES:
        _build_hotel_sheet(wb, h)

    out = tmp_path / "sample-contract.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def single_hotel_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Naama Bay Hotel"
    ws.append(["Naama Bay Hotel"])
    ws.append([])
    ws.append(["FROM", "TO", "Room", "DBL", "SGL", "CHD(2-11.99)"])
    ws.append(["01/05/2025", "31/10/2025", "Sea View", 200, 280, "Free"])
    out = tmp_path / "naama.xlsx"
    wb.save(out)
    return out
