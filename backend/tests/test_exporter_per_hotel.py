"""Tests that the exporter writes one sheet per distinct Hotel Name when
there's more than one hotel, plus a combined Hotel sheet."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.exporter import _sanitize_sheet_name, export_workbook


def test_sanitize_sheet_name_removes_invalid_chars() -> None:
    taken: set[str] = set()
    assert _sanitize_sheet_name("Naama/Bay [Resort]", taken=taken).startswith("Naama_Bay _Resort_")


def test_sanitize_sheet_name_truncates_to_31_chars() -> None:
    taken: set[str] = set()
    long = "A very long hotel name that definitely exceeds the limit"
    out = _sanitize_sheet_name(long, taken=taken)
    assert len(out) <= 31


def test_sanitize_sheet_name_deduplicates() -> None:
    taken: set[str] = set()
    a = _sanitize_sheet_name("Acrotel", taken=taken)
    b = _sanitize_sheet_name("Acrotel", taken=taken)
    assert a != b


def test_export_creates_per_hotel_sheet_when_multiple_hotels(tmp_path: Path) -> None:
    result = {
        "workbookSummary": {"sourceFile": "multi.xlsx", "inputFormat": "xlsx"},
        "dynamicColumns": {"childColumns": []},
        "hotels": [],
        "hotelRows": [
            {
                "id": "r1", "Hotel Name": "Naama Bay Hotel", "Room Name": "Sea View",
                "Start Date": "2026-04-01", "End Date": "2026-06-13",
                "Meal Plan": "Bed & Breakfast", "Currency": "EUR", "DBL": 100,
                "dynamicChildValues": {},
            },
            {
                "id": "r2", "Hotel Name": "Naama Bay Hotel", "Room Name": "Garden View",
                "Start Date": "2026-04-01", "End Date": "2026-06-13",
                "Meal Plan": "Bed & Breakfast", "Currency": "EUR", "DBL": 85,
                "dynamicChildValues": {},
            },
            {
                "id": "r3", "Hotel Name": "Mövenpick Sharm", "Room Name": "Standard",
                "Start Date": "2026-04-01", "End Date": "2026-06-13",
                "Meal Plan": "Half Board", "Currency": "EUR", "DBL": 120,
                "dynamicChildValues": {},
            },
        ],
        "extractionNotes": [],
        "validationIssues": [],
    }
    out = tmp_path / "multi.xlsx"
    export_workbook(result, output_path=out, mode="dynamic_export")
    wb = load_workbook(out, data_only=True)
    # Combined Hotel sheet + one per hotel + Extraction Notes
    assert "Hotel" in wb.sheetnames
    assert "Extraction Notes" in wb.sheetnames
    assert "Naama Bay Hotel" in wb.sheetnames
    assert "Mövenpick Sharm" in wb.sheetnames

    # Combined has all 3 rows
    assert wb["Hotel"].max_row == 4  # header + 3 rows
    # Per-hotel sheets only have their own rows
    assert wb["Naama Bay Hotel"].max_row == 3  # header + 2
    assert wb["Mövenpick Sharm"].max_row == 2  # header + 1


def test_export_keeps_single_sheet_when_only_one_hotel(tmp_path: Path) -> None:
    result = {
        "workbookSummary": {"sourceFile": "single.xlsx", "inputFormat": "xlsx"},
        "dynamicColumns": {"childColumns": []},
        "hotels": [],
        "hotelRows": [
            {
                "id": "r1", "Hotel Name": "Acrotel", "Room Name": "Double Room",
                "Start Date": "2026-04-01", "End Date": "2026-06-13",
                "Meal Plan": "Bed & Breakfast", "Currency": "EUR", "DBL": 34,
                "dynamicChildValues": {},
            },
        ],
        "extractionNotes": [],
        "validationIssues": [],
    }
    out = tmp_path / "single.xlsx"
    export_workbook(result, output_path=out, mode="dynamic_export")
    wb = load_workbook(out, data_only=True)
    # Only the combined Hotel + Extraction Notes — no extra hotel-specific sheet
    assert set(wb.sheetnames) == {"Hotel", "Extraction Notes"}
