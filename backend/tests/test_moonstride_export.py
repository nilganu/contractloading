"""Tests for the Moonstride template exporter + auto-detection."""
from __future__ import annotations

import openpyxl

from app.services.exporter import export_workbook
from app.services.moonstride_templates import (
    _band_assignment,
    _classify_band,
    _position_assignment,
    days_to_moonstride,
    detect_template,
    preview_moonstride,
)


def _row(rate_type: str, **over) -> dict:
    base = {
        "id": "r1",
        "sourceSheetOrPage": "S",
        "Hotel Name": "Hotel A",
        "Room Name": "Standard",
        "Start Date": "2025-05-01",
        "End Date": "2025-10-31",
        "Currency": "EUR",
        "Rate Type": rate_type,
        "Min Adult": 1,
        "Max Adult": 2,
        "Max Pax": 3,
        "Days": "0,1,2,3,4,5,6",
        "Meal Plan": "Bed & Breakfast",
        "Country Code ": "EG",
        "State / Province / Region": "South Sinai",
        "City / Area": "Sharm",
        "SGL": 120,
        "DBL": 80,
        "TPL": 70,
        "QDP": None,
        "Extra Bed": 25,
        "dynamicChildValues": {},
    }
    base.update(over)
    return base


def _result(rows, child_cols=None) -> dict:
    return {
        "workbookSummary": {"sourceFile": "x.xlsx", "inputFormat": "xlsx"},
        "dynamicColumns": {"childColumns": child_cols or []},
        "hotelRows": rows,
        "extractionNotes": [
            {"Source File": "x.xlsx", "Page": "1", "Category": "Taxes/service", "Note": "City tax"}
        ],
    }


def _hotel_row_dict(ws, row_idx: int) -> dict:
    headers = [c.value for c in ws[1]]
    return {
        headers[i]: ws[row_idx][i].value
        for i in range(len(headers))
        if headers[i] is not None
    }


# --------------------------------------------------------------------------
# Detection
# --------------------------------------------------------------------------
def test_detect_per_person() -> None:
    assert detect_template(_result([_row("Per Person Per Night")])) == "moonstride_ppn"
    assert detect_template(_result([_row("Per Person Per Day")])) == "moonstride_ppn"


def test_detect_per_room_defaults_to_adult_child() -> None:
    assert detect_template(_result([_row("Per Room Per Night")])) == "moonstride_prn_ac"


def test_detect_per_room_pax() -> None:
    assert (
        detect_template(_result([_row("Per Room Per Night (Pax count)")]))
        == "moonstride_prn_pax"
    )


def test_detect_empty_falls_back_to_ppn() -> None:
    assert detect_template(_result([_row("")])) == "moonstride_ppn"


def test_detect_uses_dominant_rate_type() -> None:
    rows = [_row("Per Room Per Night"), _row("Per Room Per Night"), _row("Per Person Per Night")]
    assert detect_template(_result(rows)) == "moonstride_prn_ac"


# --------------------------------------------------------------------------
# Days mask conversion
# --------------------------------------------------------------------------
def test_days_all_week() -> None:
    assert days_to_moonstride("0,1,2,3,4,5,6") == "1234567"
    assert days_to_moonstride(None) == "1234567"
    assert days_to_moonstride("") == "1234567"


def test_days_already_moonstride_passthrough() -> None:
    assert days_to_moonstride("1234567") == "1234567"
    assert days_to_moonstride("12345") == "12345"
    assert days_to_moonstride("67") == "67"


def test_days_weekdays_only() -> None:
    # internal Mon..Fri = 1,2,3,4,5 -> Moonstride 12345
    assert days_to_moonstride("1,2,3,4,5") == "12345"


def test_days_sunday_maps_to_seven() -> None:
    # internal Sunday(0) + Saturday(6) -> Moonstride 6,7 -> "67"
    assert days_to_moonstride("0,6") == "67"


def test_days_range_notation() -> None:
    assert days_to_moonstride("0 to 6") == "1234567"
    assert days_to_moonstride("1-5") == "12345"


# --------------------------------------------------------------------------
# Child band classification
# --------------------------------------------------------------------------
def test_classify_band() -> None:
    assert _classify_band(0, 1.99) == "baby"
    assert _classify_band(0, 2) == "baby"
    assert _classify_band(2, 11.99) == "child"
    assert _classify_band(3, 11) == "child"
    assert _classify_band(12, 15.99) == "teen"
    assert _classify_band(None, None) == "child"


def test_band_assignment_primary_secondary() -> None:
    cols = [
        {"key": "CHD1(2-6)", "ageFrom": 2, "ageTo": 6},
        {"key": "CHD2(7-11.99)", "ageFrom": 7, "ageTo": 11.99},
        {"key": "CHD(0-1.99)", "ageFrom": 0, "ageTo": 1.99},
    ]
    bands = _band_assignment(cols)
    assert bands["child"]["primary"] == "CHD1(2-6)"
    assert bands["child"]["secondary"] == "CHD2(7-11.99)"
    assert bands["child"]["age_from"] == 2
    assert bands["child"]["age_to"] == 11.99
    assert bands["baby"]["primary"] == "CHD(0-1.99)"
    assert bands["teen"]["primary"] is None


def test_position_assignment_uses_child_position() -> None:
    cols = [
        {"key": "CHD1(0.1-11.99)", "ageFrom": 0.1, "ageTo": 11.99, "childPosition": "first_child"},
        {"key": "CHD2(2-11.99)", "ageFrom": 2, "ageTo": 11.99, "childPosition": "second_child"},
        {"key": "CHD3(2-11.99)", "ageFrom": 2, "ageTo": 11.99, "childPosition": "third_child"},
    ]
    pos = _position_assignment(cols)
    assert pos[1]["key"] == "CHD1(0.1-11.99)"
    assert pos[1]["age_from"] == 0.1
    assert pos[1]["age_to"] == 11.99
    assert pos[2]["key"] == "CHD2(2-11.99)"
    assert pos[3]["key"] == "CHD3(2-11.99)"


def test_position_assignment_falls_back_to_order() -> None:
    # No childPosition -> sequential by contract order.
    cols = [
        {"key": "CHD(0-1.99)", "ageFrom": 0, "ageTo": 1.99},
        {"key": "CHD(2-11.99)", "ageFrom": 2, "ageTo": 11.99},
    ]
    pos = _position_assignment(cols)
    assert pos[1]["key"] == "CHD(0-1.99)"
    assert pos[2]["key"] == "CHD(2-11.99)"
    assert 3 not in pos


# --------------------------------------------------------------------------
# Full export — Per Person Per Night
# --------------------------------------------------------------------------
def test_export_ppn_layout(tmp_path) -> None:
    child_cols = [
        {"key": "CHD1(0.1-11.99)", "ageFrom": 0.1, "ageTo": 11.99,
         "childPosition": "first_child", "valueType": "amount"},
        {"key": "CHD2(2-11.99)", "ageFrom": 2, "ageTo": 11.99,
         "childPosition": "second_child", "valueType": "amount"},
        {"key": "CHD3(2-11.99)", "ageFrom": 2, "ageTo": 11.99,
         "childPosition": "third_child", "valueType": "not_applicable"},
    ]
    rows = [
        _row(
            "Per Person Per Night",
            dynamicChildValues={"CHD1(0.1-11.99)": 20, "CHD2(2-11.99)": 15, "CHD3(2-11.99)": None},
        )
    ]
    out = tmp_path / "ppn.xlsx"
    export_workbook(_result(rows, child_cols), output_path=str(out), mode="moonstride_auto")

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Hotel", "MasterData", "Extraction Notes"]
    ws = wb["Hotel"]
    rd = _hotel_row_dict(ws, 2)

    assert rd["Rate Type"] == "Per Person Per Night"
    assert rd["Days"] == "1234567"
    assert rd["Room Name"] == "Standard"
    assert rd["Country Code"] == "EG"
    assert rd["County / State / Province"] == "South Sinai"
    # adult rate mapping
    assert rd["Adult 1 (SGL)"] == 120
    assert rd["Adult 2 (DBL)"] == 80
    assert rd["Adult 3 (TPL)"] == 70
    assert rd["Adult 4 (QUD)"] is None
    assert rd["Extra Adult"] == 25
    # defaults
    assert rd["Status"] == "Open"
    assert rd["Customer Price Currency"] == "EUR"
    assert rd["Check-In"] == "00:00"
    # per-position child columns (Price / Age Min / Age Max), mirroring the contract
    assert rd["1st Child Price"] == 20
    assert rd["1st Child Age Min"] == 0.1
    assert rd["1st Child Age Max"] == 11.99
    assert rd["2nd Child Price"] == 15
    assert rd["2nd Child Age Min"] == 2
    assert rd["2nd Child Age Max"] == 11.99
    assert rd["3rd Child Price"] is None
    assert rd["3rd Child Age Min"] == 2
    assert rd["3rd Child Age Max"] == 11.99
    # the template's single age columns are left blank — position columns own this
    assert rd["Child Age"] is None
    assert rd["Infant Age"] is None


def test_export_ppn_header_matches_template(tmp_path) -> None:
    out = tmp_path / "ppn.xlsx"
    export_workbook(
        _result([_row("Per Person Per Night")]), output_path=str(out), mode="moonstride_ppn"
    )
    produced = [c.value for c in openpyxl.load_workbook(out)["Hotel"][1] if c.value is not None]

    template = openpyxl.load_workbook(
        "app/templates/moonstride/per_person_per_night.xlsx", read_only=True
    )["Hotel"]
    tmpl_headers = [c for c in next(template.iter_rows(values_only=True)) if c is not None]

    # Produced headers must start with the exact template headers, then append
    # the per-position child columns (Price / Age Min / Age Max x 3).
    assert produced[: len(tmpl_headers)] == tmpl_headers
    assert produced[len(tmpl_headers):] == [
        "1st Child Price", "1st Child Age Min", "1st Child Age Max",
        "2nd Child Price", "2nd Child Age Min", "2nd Child Age Max",
        "3rd Child Price", "3rd Child Age Min", "3rd Child Age Max",
    ]


# --------------------------------------------------------------------------
# Full export — Per Room Per Night (Pax count)
# --------------------------------------------------------------------------
def test_export_pax_layout(tmp_path) -> None:
    rows = [_row("Per Room Per Night (Pax count)", QDP=180, Days="1,2,3,4,5")]
    out = tmp_path / "pax.xlsx"
    export_workbook(_result(rows), output_path=str(out), mode="moonstride_auto")

    ws = openpyxl.load_workbook(out)["Hotel"]
    rd = _hotel_row_dict(ws, 2)
    assert rd["Rate Type"] == "Per Room Per Night (Pax count)"
    assert rd["Days"] == "12345"
    assert rd["1 Pax"] == 120
    assert rd["2 Pax"] == 80
    assert rd["3 Pax"] == 70
    assert rd["4 Pax"] == 180
    assert rd["5 Pax"] is None
    assert rd["Extra Adult"] == 25


def test_preview_matches_export_structure() -> None:
    child_cols = [
        {"key": "CHD1(0.1-11.99)", "ageFrom": 0.1, "ageTo": 11.99,
         "childPosition": "first_child", "valueType": "amount"},
    ]
    rows = [_row("Per Person Per Night", dynamicChildValues={"CHD1(0.1-11.99)": 20})]
    preview = preview_moonstride(_result(rows, child_cols))

    assert preview["templateId"] == "moonstride_ppn"
    assert preview["rateType"] == "Per Person Per Night"
    assert preview["headers"][-9:] == [
        "1st Child Price", "1st Child Age Min", "1st Child Age Max",
        "2nd Child Price", "2nd Child Age Min", "2nd Child Age Max",
        "3rd Child Price", "3rd Child Age Min", "3rd Child Age Max",
    ]
    assert len(preview["rows"]) == 1
    r0 = preview["rows"][0]
    assert r0["Rate Type"] == "Per Person Per Night"
    assert r0["Days"] == "1234567"
    assert r0["Adult 1 (SGL)"] == 120
    assert r0["1st Child Price"] == 20
    assert r0["1st Child Age Min"] == 0.1
    # dates are JSON-safe ISO strings
    assert r0["Start Date"] == "2025-05-01"


def test_export_extraction_notes_sheet(tmp_path) -> None:
    out = tmp_path / "ppn.xlsx"
    export_workbook(
        _result([_row("Per Person Per Night")]), output_path=str(out), mode="moonstride_auto"
    )
    ws = openpyxl.load_workbook(out)["Extraction Notes"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Source File", "Page", "Category", "Note")
    assert rows[1][2] == "Taxes/service"
