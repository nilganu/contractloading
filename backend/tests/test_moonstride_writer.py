"""Tests for the deterministic Moonstride hotel-file writer.

The writer used to auto-append per-position child columns (1st/2nd/3rd
Child Price / Age Min / Age Max). Per Jun 2026 user rule the hotel
file no longer carries any child-band or per-position child columns —
child policy lives in the supplement file. This test asserts those
columns are absent from the written output.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import openpyxl

from app.services.moonstride_templates import write_raw_rows


def _row() -> Dict[str, Any]:
    return {
        "Hotel Name": "Test Hotel",
        "Hotel Code": "100001",
        "Room Name": "Standard",
        "Season": "Summer",
        "Start Date": "2026-05-01",
        "End Date": "2026-09-30",
        "Meal Plan": "Bed and Breakfast",
        "Adult 1 (SGL)": 100,
        "Adult 2 (DBL)": 160,
        "Currency": "EUR",
        "Status": "Open",
        "Days": "1234567",
    }


def _all_headers(path: Path) -> List[str]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Hotel"] if "Hotel" in wb.sheetnames else wb.active
    return [
        str(ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]


def test_write_raw_rows_drops_child_band_columns(tmp_path: Path) -> None:
    out = tmp_path / "h.xlsx"
    write_raw_rows([_row()], "moonstride_ppn", out)
    headers = _all_headers(out)
    for forbidden in (
        "Baby 1 (0-1)", "Child 1 (2-12)", "Teen 1 (12-17)",
        "Multi Infant (0-1)", "Extra Child (2-12)", "Extra Teen (12-17)",
        "1st Child Price", "1st Child Age Min", "1st Child Age Max",
        "2nd Child Price", "2nd Child Age Min", "2nd Child Age Max",
        "3rd Child Price", "3rd Child Age Min", "3rd Child Age Max",
    ):
        assert forbidden not in headers, (
            f"hotel template still carries forbidden header {forbidden!r}"
        )


def test_write_raw_rows_preserves_hotel_metadata_columns(tmp_path: Path) -> None:
    """Sanity: removing child columns doesn't accidentally drop other
    metadata headers the importer expects."""
    out = tmp_path / "h.xlsx"
    write_raw_rows([_row()], "moonstride_ppn", out)
    headers = _all_headers(out)
    for expected in (
        "Hotel Name", "Hotel Code",
        "Address Line 1", "Address Line 2", "Postal Code", "Country Code",
        "City / Area", "Phone Number", "Email Address", "Hotel Website",
        "Latitude", "Longitude",
        "Room Name", "Season", "Start Date", "End Date",
        "Adult 1 (SGL)", "Adult 2 (DBL)",
        "Currency", "Status", "Days",
    ):
        assert expected in headers, (
            f"hotel template lost expected header {expected!r}"
        )


def test_write_raw_rows_recolours_ai_filled_cells(tmp_path: Path) -> None:
    """The writer must consume the row's ``_ai_fields`` side-channel
    (set of header names) and re-paint those cells with the AI-marker
    font AFTER the standard colour-strip. The ``_ai_fields`` key
    itself must never be written to a column."""
    out = tmp_path / "h.xlsx"
    row = _row()
    row["Address Line 1"] = "Naama Bay, Sharm El Sheikh"
    row["Phone Number"] = "+20 69 360 0100"
    row["_ai_fields"] = {"Address Line 1", "Phone Number"}
    write_raw_rows([row], "moonstride_ppn", out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Hotel"] if "Hotel" in wb.sheetnames else wb.active
    headers = [
        str(ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    # Side-channel key must NOT have leaked as a column header.
    assert "_ai_fields" not in headers
    idx = {h: i + 1 for i, h in enumerate(headers)}
    # AI cells should be italic + non-black colour; non-AI cells stay
    # plain black non-italic.
    ai_cell = ws.cell(row=2, column=idx["Address Line 1"])
    assert ai_cell.value == "Naama Bay, Sharm El Sheikh"
    assert ai_cell.font.italic is True
    assert (ai_cell.font.color.rgb or "").upper() != "FF000000"
    plain_cell = ws.cell(row=2, column=idx["Hotel Name"])
    assert plain_cell.font.italic is False or plain_cell.font.italic is None
    assert (plain_cell.font.color.rgb or "FF000000").upper() == "FF000000"


def test_write_raw_rows_writes_priced_values(tmp_path: Path) -> None:
    out = tmp_path / "h.xlsx"
    write_raw_rows([_row()], "moonstride_ppn", out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Hotel"] if "Hotel" in wb.sheetnames else wb.active
    headers = [
        str(ws.cell(row=1, column=c).value or "").strip()
        for c in range(1, ws.max_column + 1)
    ]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    assert ws.max_row == 2
    assert ws.cell(row=2, column=idx["Adult 1 (SGL)"]).value == 100
    assert ws.cell(row=2, column=idx["Adult 2 (DBL)"]).value == 160
    assert ws.cell(row=2, column=idx["Hotel Code"]).value == "100001"
