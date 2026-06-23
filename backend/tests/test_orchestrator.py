"""Unit tests for the Phase 2 orchestrator pieces — outline, splitters,
and aggregation. The actual LLM-calling orchestrate_extraction is not
exercised here; it's covered by the manual end-to-end smoke (live API).
"""
from __future__ import annotations

import openpyxl
from openpyxl import Workbook

from app.extraction.canonical import (
    ContractExtraction, HotelExtraction, HotelMetadata, MealPlanEntry,
    Rate, RoomType, Season, Supplement,
)
from app.extraction.orchestrator import (
    _aggregate, _build_missing_rates_directive,
    _build_missing_supplements_directive, _dedup_canonical,
    _dedup_supplements, _merge_new_rates, _merge_new_supplements,
)
from app.extraction.verifier import (
    VerifierFinding, VerifierReport, collect_missing_supplement_names,
)
from app.extraction.outline import outline_excel_locally
from app.extraction.splitters import (
    is_index_sheet,
    list_excel_sheets,
    parse_page_hint,
    split_excel_sheet,
    split_for_hotel,
)


# --- index-sheet heuristic ---------------------------------------------------


def test_is_index_sheet_recognises_common_index_names() -> None:
    for name in [
        "Hotel List", "hotel list", "  Hotels  ", "Index", "Summary",
        "Contents", "Table of Contents", "TOC", "Overview", "Cover",
        "Legend", "Info", "Notes", "Terms", "Masterdata", "Sheet1",
    ]:
        assert is_index_sheet(name), f"expected '{name}' to be an index sheet"


def test_is_index_sheet_keeps_actual_hotel_names() -> None:
    for name in [
        "Barceló Tiran Sharm Resort", "Charmillion Sea Life",
        "Atrium Prestige", "LUX South Ari Atoll",
    ]:
        assert not is_index_sheet(name), f"expected '{name}' to be kept"


# --- Excel listing + outline -------------------------------------------------


def _multi_sheet_workbook(tmp_path, sheet_names):
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheet_names:
        ws = wb.create_sheet(name)
        ws["A1"] = f"Hello {name}"
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return path


def test_list_excel_sheets_returns_all(tmp_path) -> None:
    path = _multi_sheet_workbook(tmp_path, ["Hotel A", "Hotel B"])
    assert list_excel_sheets(path) == ["Hotel A", "Hotel B"]


def test_outline_excel_locally_skips_index_sheets(tmp_path) -> None:
    path = _multi_sheet_workbook(
        tmp_path, ["Hotel List", "Barceló Tiran", "Charmillion Sea Life", "Notes"]
    )
    outline = outline_excel_locally(path)
    assert outline.is_multi_hotel
    names = [h.name for h in outline.hotels]
    assert names == ["Barceló Tiran", "Charmillion Sea Life"]
    # source_hint mirrors the sheet name
    assert outline.hotels[0].source_hint == "Sheet:Barceló Tiran"


def test_outline_excel_locally_single_hotel(tmp_path) -> None:
    path = _multi_sheet_workbook(tmp_path, ["Atrium Prestige"])
    outline = outline_excel_locally(path)
    assert not outline.is_multi_hotel
    assert [h.name for h in outline.hotels] == ["Atrium Prestige"]


# --- split_excel_sheet -------------------------------------------------------


def test_split_excel_sheet_round_trip(tmp_path) -> None:
    src = _multi_sheet_workbook(
        tmp_path, ["Hotel A", "Hotel B", "Hotel C"]
    )
    dest = tmp_path / "out.xlsx"
    split_excel_sheet(src, "Hotel B", dest)
    out_wb = openpyxl.load_workbook(dest)
    assert out_wb.sheetnames == ["Hotel B"]
    assert out_wb["Hotel B"]["A1"].value == "Hello Hotel B"


# --- split_for_hotel ---------------------------------------------------------


def test_split_for_hotel_excel_match_by_exact_sheet(tmp_path) -> None:
    src = _multi_sheet_workbook(
        tmp_path, ["Hotel List", "Barceló Tiran Sharm Resort", "Charmillion Sea Life"]
    )
    sub_dir = tmp_path / "subs"
    out, mode = split_for_hotel(
        src, "Barceló Tiran Sharm Resort",
        "Sheet:Barceló Tiran Sharm Resort", sub_dir,
    )
    assert mode == "excel_sheet"
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Barceló Tiran Sharm Resort"]


def test_split_for_hotel_excel_falls_back_to_whole_when_sheet_unknown(tmp_path) -> None:
    src = _multi_sheet_workbook(tmp_path, ["Some Hotel"])
    sub_dir = tmp_path / "subs"
    out, mode = split_for_hotel(src, "Missing Hotel", None, sub_dir)
    # Fallback copies the whole file.
    assert mode == "whole"
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Some Hotel"]


# --- PDF page-hint parser ----------------------------------------------------


def test_parse_page_hint_range() -> None:
    assert parse_page_hint("Pages 1-3", 10) == [1, 2, 3]


def test_parse_page_hint_single() -> None:
    assert parse_page_hint("Page 4", 10) == [4]


def test_parse_page_hint_comma_list() -> None:
    assert parse_page_hint("2, 5, 7", 10) == [2, 5, 7]


def test_parse_page_hint_empty_falls_back_to_all() -> None:
    assert parse_page_hint(None, 5) == [1, 2, 3, 4, 5]
    assert parse_page_hint("", 5) == [1, 2, 3, 4, 5]


def test_parse_page_hint_clamps_to_document_length() -> None:
    assert parse_page_hint("Pages 8-20", 10) == [8, 9, 10]


# --- aggregation -------------------------------------------------------------


def _single_hotel(name: str, supp_count: int = 0) -> ContractExtraction:
    h = HotelExtraction(
        metadata=HotelMetadata(name=name, currency="EUR"),
        rooms=[RoomType(name="Standard", max_pax=2)],
        seasons=[Season(label="S", start_date="2025-05-01", end_date="2025-09-30")],
        meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
        rates=[Rate(room_name="Standard", season_label="S", meal_code="BB", dbl=100)],
        child_policy=[],
    )
    supps = [
        Supplement(
            name=f"S{i}", hotel_name=name, kind="meal_upgrade",
            charge_type="Per Person Per Night", calculation_method="Standard",
            traveler_type="Adult",
        )
        for i in range(supp_count)
    ]
    return ContractExtraction(
        source_filename="x.xlsx", is_multi_hotel=False,
        detected_rate_type="Per Person Per Night",
        hotels=[h], supplements=supps, notes=[],
    )


def test_aggregate_concatenates_hotels_and_supplements() -> None:
    a = _single_hotel("Alpha", supp_count=2)
    b = _single_hotel("Beta", supp_count=1)
    merged = _aggregate([a, b], source_filename="contract.xlsx")
    assert merged.is_multi_hotel
    assert {h.metadata.name for h in merged.hotels} == {"Alpha", "Beta"}
    assert len(merged.supplements) == 3
    assert merged.source_filename == "contract.xlsx"


def test_aggregate_dedups_hotels_by_name() -> None:
    """A hotel reported twice (e.g. once on its sheet and once in the
    'Hotel List' index) should appear only once in the merged output."""
    a = _single_hotel("Alpha")
    a2 = _single_hotel("alpha")  # case insensitive dedup
    merged = _aggregate([a, a2], source_filename="x")
    assert len(merged.hotels) == 1


def test_aggregate_picks_modal_rate_type() -> None:
    a = _single_hotel("A")
    b = _single_hotel("B")
    c = _single_hotel("C")
    # Override one to a different rate type
    c.detected_rate_type = "Per Room Per Night"
    merged = _aggregate([a, b, c], source_filename="x")
    assert merged.detected_rate_type == "Per Person Per Night"  # 2 of 3


def test_aggregate_rate_type_vote_weighted_by_priced_rates() -> None:
    """A hotel with 1 priced row classified as PRN should NOT outvote two
    hotels carrying 50 priced rows each classified as PPN."""
    a = _single_hotel("A")
    a.hotels[0].rates = [
        Rate(room_name="Standard", season_label="S", meal_code="BB", dbl=100)
        for _ in range(50)
    ]
    b = _single_hotel("B")
    b.hotels[0].rates = [
        Rate(room_name="Standard", season_label="S", meal_code="BB", dbl=100)
        for _ in range(50)
    ]
    # Tiny hotel that misclassifies — should be drowned out.
    c = _single_hotel("C")
    c.detected_rate_type = "Per Room Per Night"
    merged = _aggregate([a, b, c], source_filename="x")
    assert merged.detected_rate_type == "Per Person Per Night"


# --- Phase 3: retry directive + merge -----------------------------------------


def test_build_missing_rates_directive_lists_triples_verbatim() -> None:
    triples = [("Room A", "Summer", "BB"), ("Room A", "Winter", "BB")]
    s = _build_missing_rates_directive("Some Hotel", triples)
    assert "Some Hotel" in s
    assert "'Room A'" in s and "'Summer'" in s and "'Winter'" in s
    assert "'BB'" in s


def test_build_missing_rates_directive_overflow_summarises() -> None:
    triples = [("R" + str(i), "S", "BB") for i in range(80)]
    s = _build_missing_rates_directive("H", triples)
    assert "Plus" in s and "20 more" in s  # 80 - 60 cap = 20


def test_merge_new_rates_appends_only_missing_priced_rows() -> None:
    base = _single_hotel("Test")
    base.hotels[0].rates = [
        Rate(room_name="A", season_label="S1", meal_code="BB", dbl=100),
    ]
    # Patch: tries to add an A/S2/BB (priced, missing) and an A/S1/BB (already present).
    patch = ContractExtraction(
        source_filename="p.xlsx", is_multi_hotel=False,
        detected_rate_type="Per Person Per Night",
        hotels=[HotelExtraction(
            metadata=HotelMetadata(name="Test", currency="EUR"),
            rooms=[RoomType(name="A", max_pax=2)],
            seasons=[Season(label="S2", start_date="2025-05-01", end_date="2025-09-30")],
            meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
            rates=[
                Rate(room_name="A", season_label="S2", meal_code="BB", dbl=120),
                Rate(room_name="A", season_label="S1", meal_code="BB", dbl=999),  # dup
                Rate(room_name="A", season_label="S3", meal_code="BB", dbl=50),   # not in missing
                Rate(room_name="A", season_label="S2", meal_code="BB"),           # null prices
            ],
            child_policy=[],
        )],
        supplements=[], notes=[],
    )
    target = base.hotels[0]
    _merge_new_rates(target, patch, [("A", "S2", "BB"), ("A", "S3", "BB")])
    # Only the priced A/S2/BB should land; the dup should NOT overwrite the
    # existing A/S1/BB; A/S3/BB had a row but was priced so it lands too.
    keys = [(r.room_name, r.season_label, r.meal_code, r.dbl) for r in target.rates]
    assert ("A", "S1", "BB", 100) in keys
    assert ("A", "S2", "BB", 120) in keys
    assert ("A", "S3", "BB", 50) in keys
    assert ("A", "S1", "BB", 999) not in keys  # original protected


def test_dedup_canonical_collapses_duplicate_seasons() -> None:
    h = HotelExtraction(
        metadata=HotelMetadata(name="Test", currency="EUR"),
        rooms=[RoomType(name="A", max_pax=2)],
        seasons=[
            Season(label="S1", start_date="2025-05-01", end_date="2025-09-30"),
            Season(label="S1", start_date="2025-05-01", end_date="2025-09-30"),  # dup
            Season(label="S2", start_date="2025-10-01", end_date="2026-03-31"),
        ],
        meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
        rates=[],
        child_policy=[],
    )
    _dedup_canonical(h)
    assert len(h.seasons) == 2
    assert {s.label for s in h.seasons} == {"S1", "S2"}


def test_dedup_canonical_picks_most_filled_rate() -> None:
    """When multiple rates share (room, season, meal), keep the row with
    the most non-null occupancy prices."""
    sparse = Rate(room_name="A", season_label="S", meal_code="BB", dbl=100)
    full = Rate(
        room_name="A", season_label="S", meal_code="BB",
        sgl=80, dbl=100, tpl=130, qdp=160,
    )
    h = HotelExtraction(
        metadata=HotelMetadata(name="Test", currency="EUR"),
        rooms=[RoomType(name="A", max_pax=2)],
        seasons=[Season(label="S", start_date="2025-05-01", end_date="2025-09-30")],
        meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
        rates=[sparse, full],  # sparse comes first; full should win
        child_policy=[],
    )
    _dedup_canonical(h)
    assert len(h.rates) == 1
    assert h.rates[0].sgl == 80 and h.rates[0].tpl == 130


def test_dedup_canonical_no_change_when_already_clean() -> None:
    h = HotelExtraction(
        metadata=HotelMetadata(name="Test", currency="EUR"),
        rooms=[RoomType(name="A", max_pax=2)],
        seasons=[Season(label="S", start_date="2025-05-01", end_date="2025-09-30")],
        meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
        rates=[Rate(room_name="A", season_label="S", meal_code="BB", dbl=100)],
        child_policy=[],
    )
    _dedup_canonical(h)
    assert len(h.seasons) == 1 and len(h.rates) == 1


def test_dedup_canonical_collapses_duplicate_child_bands() -> None:
    """Source contracts often list the SAME band in both the inline
    child columns on the rate grid AND in the explicit 'Children Policy'
    section. The LLM emits both. Dedup by (position, age range,
    value type, value)."""
    from app.extraction.canonical import ChildAgeBand
    h = HotelExtraction(
        metadata=HotelMetadata(name="Test", currency="EUR"),
        rooms=[RoomType(name="A", max_pax=4)],
        seasons=[Season(label="S", start_date="2025-05-01", end_date="2025-09-30")],
        meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
        rates=[Rate(room_name="A", season_label="S", meal_code="BB", dbl=40)],
        child_policy=[
            ChildAgeBand(position="first_child", age_from=0, age_to=5.99,
                         value_type="discount_percentage", value=100),
            ChildAgeBand(position="first_child", age_from=0, age_to=5.99,
                         value_type="discount_percentage", value=100),  # dupe
            ChildAgeBand(position="first_child", age_from=6, age_to=10.99,
                         value_type="discount_percentage", value=50),
            ChildAgeBand(position="second_child", age_from=6, age_to=10.99,
                         value_type="discount_percentage", value=25),
            ChildAgeBand(position="second_child", age_from=6, age_to=10.99,
                         value_type="discount_percentage", value=25),  # dupe
        ],
    )
    _dedup_canonical(h)
    assert len(h.child_policy) == 3
    # Genuine variants kept; pure duplicates collapsed.
    signatures = {
        (b.position, b.age_from, b.age_to, b.value_type, b.value)
        for b in h.child_policy
    }
    assert ("first_child", 0, 5.99, "discount_percentage", 100) in signatures
    assert ("first_child", 6, 10.99, "discount_percentage", 50) in signatures
    assert ("second_child", 6, 10.99, "discount_percentage", 25) in signatures


# --- Phase 4: verifier helpers + merge ----------------------------------------


def test_collect_missing_supplement_names_dedups_and_strips() -> None:
    report = VerifierReport(
        hotel_name="CGR",
        findings=[
            VerifierFinding(
                finding_kind="MISSING_SUPPLEMENT", severity="error",
                field_path="supplements[name='X-Mass Italian']",
                observation="X-Mass Gala Italian variant absent",
                missing_supplement_name=" X-Mass Italian ",
            ),
            VerifierFinding(
                finding_kind="MISSING_SUPPLEMENT", severity="warning",
                field_path="supplements[name='X-Mass Italian']",
                observation="duplicate",
                missing_supplement_name="x-mass italian",  # same after lower+strip
            ),
            VerifierFinding(
                finding_kind="WRONG_VALUE", severity="error",
                field_path="supplements[0].supplier_cost",
                observation="Cost is 50, contract says 60",
                missing_supplement_name=None,  # WRONG_VALUE doesn't fill this
            ),
        ],
    )
    names = collect_missing_supplement_names(report)
    assert names == ["X-Mass Italian"]


def test_build_missing_supplements_directive_includes_names() -> None:
    s = _build_missing_supplements_directive(
        "Charmillion Garden Resort",
        ["X-Mass Gala — Italian Market", "NY Gala — Italian Market"],
    )
    assert "Charmillion Garden Resort" in s
    assert "X-Mass Gala — Italian Market" in s
    assert "NY Gala — Italian Market" in s


def test_merge_new_supplements_dedups_and_appends() -> None:
    base = _single_hotel("CGR", supp_count=1)  # 1 existing supplement
    # Patch contains: same supplement (skipped) + new one (added) + one tagged for OTHER hotel.
    patch = ContractExtraction(
        source_filename="p.xlsx", is_multi_hotel=False,
        detected_rate_type="Per Person Per Night",
        hotels=base.hotels,
        supplements=[
            Supplement(
                name="S0", hotel_name="CGR", kind="meal_upgrade",
                charge_type="Per Person Per Night",
                calculation_method="Standard", traveler_type="Adult",
            ),  # duplicate of existing
            Supplement(
                name="New gala", hotel_name="CGR", kind="gala_dinner",
                charge_type="Per Person Per Night",
                calculation_method="Standard", traveler_type="Adult",
                supplier_cost=75,
            ),
            Supplement(
                name="Wrong hotel supp", hotel_name="OtherHotel",
                kind="gala_dinner",
                charge_type="Per Person Per Night",
                calculation_method="Standard", traveler_type="Adult",
            ),  # patch tagged it for OtherHotel, but the directive targeted CGR — force CGR
        ],
        notes=[],
    )
    _merge_new_supplements(base, patch, "CGR")
    names = [s.name for s in base.supplements]
    assert "New gala" in names
    assert names.count("S0") == 1  # not duplicated
    # The misrouted "Wrong hotel supp" gets adopted under CGR per our merge rule.
    assert "Wrong hotel supp" in names
    assert all(s.hotel_name == "CGR" for s in base.supplements if s.name in ("New gala", "Wrong hotel supp"))


# --- _dedup_supplements: name-insensitive identity ----------------------------


def _gala(name: str, hotel: str = "CGR", cost: float = 50, trav: str = "Adult",
          start: str = "2025-12-24", end: str = "2025-12-24") -> Supplement:
    return Supplement(
        name=name, hotel_name=hotel, kind="gala_dinner",
        charge_type="Per Person Per Night", calculation_method="Standard",
        traveler_type=trav, supplier_cost=cost, customer_price=cost,
        start_date=start, end_date=end,
    )


def test_dedup_supplements_collapses_renamed_duplicates() -> None:
    """LLM emits the SAME supplement twice with rephrased names —
    "X- Mass Gala Dinner 24.12.25 Obligatory" vs "X-Mass Gala Dinner —
    24.12.25" — the dedup must collapse them."""
    ext = ContractExtraction(
        source_filename="x.xlsx", is_multi_hotel=False,
        detected_rate_type="Per Person Per Night",
        hotels=[_single_hotel("CGR").hotels[0]],
        supplements=[
            _gala("X- Mass Gala Dinner 24.12.25 Obligatory"),
            _gala("X-Mass Gala Dinner — 24.12.25"),  # rephrased
            _gala("New Year Gala Dinner 31.12.25 Obligatory", cost=110,
                  start="2025-12-31", end="2025-12-31"),
        ],
        notes=[],
    )
    _dedup_supplements(ext)
    assert len(ext.supplements) == 2  # one X-Mass, one NY


def test_dedup_supplements_keeps_genuine_variants_by_date() -> None:
    """Same name + same cost but DIFFERENT dates → genuinely different
    supplements; do NOT collapse."""
    ext = ContractExtraction(
        source_filename="x.xlsx", is_multi_hotel=False,
        detected_rate_type="Per Person Per Night",
        hotels=[_single_hotel("CGR").hotels[0]],
        supplements=[
            _gala("Gala", start="2025-12-24", end="2025-12-24"),
            _gala("Gala", start="2025-12-31", end="2025-12-31"),  # different date
        ],
        notes=[],
    )
    _dedup_supplements(ext)
    assert len(ext.supplements) == 2


def test_dedup_supplements_collapses_null_vs_filled_cost_variants() -> None:
    """LLM emits the SAME supplement twice — once with supplier_cost
    set and once with customer_price set. The cost key normalises
    to whichever is non-null so both collapse into one row, keeping
    the most complete variant."""
    s1 = Supplement(
        name="Early Booking", hotel_name="CGR", kind="special_offer",
        charge_type="Per Person Per Night", calculation_method="Standard",
        traveler_type="Traveller",
        supplier_cost=15, customer_price=None,
    )
    s2 = Supplement(
        name="Early Booking Discount 15%", hotel_name="CGR",
        kind="special_offer",
        charge_type="Per Person Per Night", calculation_method="Standard",
        traveler_type="Traveller",
        supplier_cost=15, customer_price=15,  # more complete
    )
    ext = ContractExtraction(
        source_filename="x.xlsx", is_multi_hotel=False,
        detected_rate_type="Per Person Per Night",
        hotels=[_single_hotel("CGR").hotels[0]],
        supplements=[s1, s2],
        notes=[],
    )
    _dedup_supplements(ext)
    assert len(ext.supplements) == 1
    # The more complete entry won.
    kept = ext.supplements[0]
    assert kept.customer_price == 15
    assert kept.name == "Early Booking Discount 15%"


def test_dedup_canonical_subsumes_contained_season() -> None:
    """The Acrotel "01.04-13.06 & 15.09-31.10" combined period
    contains the "15.09-31.10" half. After dedup the shorter
    contained season is dropped."""
    h = HotelExtraction(
        metadata=HotelMetadata(name="Acrotel", currency="EUR"),
        rooms=[RoomType(name="Double", max_pax=2)],
        seasons=[
            Season(label="01/04-13/06 & 15/09-31/10",
                   start_date="2026-04-01", end_date="2026-10-31"),
            Season(label="15/09-31/10",
                   start_date="2026-09-15", end_date="2026-10-31"),  # contained
            Season(label="14/06-31/08",
                   start_date="2026-06-14", end_date="2026-08-31"),  # disjoint
        ],
        meal_plans=[MealPlanEntry(code="BB", canonical="Bed and Breakfast")],
        rates=[], child_policy=[],
    )
    _dedup_canonical(h)
    labels = [s.label for s in h.seasons]
    assert "01/04-13/06 & 15/09-31/10" in labels
    assert "14/06-31/08" in labels
    assert "15/09-31/10" not in labels


def test_dedup_supplements_keeps_per_hotel_separation() -> None:
    """The same gala at two different hotels stays as two rows."""
    a = _single_hotel("A").hotels[0]
    b = _single_hotel("B").hotels[0]
    ext = ContractExtraction(
        source_filename="x.xlsx", is_multi_hotel=True,
        detected_rate_type="Per Person Per Night",
        hotels=[a, b],
        supplements=[
            _gala("Gala", hotel="A"),
            _gala("Gala — renamed", hotel="A"),  # dupe at same hotel
            _gala("Gala", hotel="B"),            # legitimate, different hotel
        ],
        notes=[],
    )
    _dedup_supplements(ext)
    assert len(ext.supplements) == 2  # one for A, one for B
