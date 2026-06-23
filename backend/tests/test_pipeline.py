"""End-to-end backend pipeline test using the stub LLM extractor."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.ir_builder import build_ir_from_excel
from app.services.normalizer import normalize_result
from app.services.parsers.excel import parse_excel
from app.services.stub_extractor import stub_extract
from app.services.validator import validate_result
from app.services.exporter import export_workbook


def _run_pipeline(xlsx_path: Path, mode: str = "dynamic_export") -> dict:
    parsed = parse_excel(xlsx_path)
    ir = build_ir_from_excel(parsed)
    options = {
        "supplierDefault": "Test Supplier",
        "countryDefault": "EG",
        "currencyDefault": "EUR",
        "statusDefault": "Open",
        "childColumnMode": mode,
        "preserveChildPositions": True,
        "extractionMode": "text_only",
    }
    raw = stub_extract(ir, options)
    norm = normalize_result(raw, options, parsed["source_file"])
    norm["validationIssues"] = validate_result(norm, options)
    return norm


def test_pipeline_produces_hotels_per_sheet(sample_xlsx: Path) -> None:
    result = _run_pipeline(sample_xlsx)
    # All 3 hotel sheets should be processed, Hotel List should be index_reference
    assert "Hotel List" not in result["workbookSummary"]["hotelSheets"]
    assert any("Hotel List" in s for s in result["workbookSummary"]["indexSheets"])
    assert len(result["workbookSummary"]["hotelSheets"]) >= 3


def test_pipeline_dynamic_child_columns(sample_xlsx: Path) -> None:
    result = _run_pipeline(sample_xlsx)
    keys = [c["key"] for c in result["dynamicColumns"]["childColumns"]]
    # Synthetic contract uses CHD(2-11.99) and CHD(0-1.99)
    assert any("2-11.99" in k for k in keys)
    assert any("0-1.99" in k for k in keys)


def test_pipeline_creates_hotel_rows(sample_xlsx: Path) -> None:
    result = _run_pipeline(sample_xlsx)
    assert len(result["hotelRows"]) > 0
    sample = result["hotelRows"][0]
    assert "Hotel Name" in sample
    assert "Start Date" in sample
    assert "End Date" in sample


def test_pipeline_extraction_notes_for_policies(sample_xlsx: Path) -> None:
    result = _run_pipeline(sample_xlsx)
    cats = {n["Category"] for n in result["extractionNotes"]}
    assert "Cancellation" in cats
    assert "Gala dinner" in cats


def test_pipeline_validation_includes_errors_when_required_missing(sample_xlsx: Path) -> None:
    # Stub extractor leaves Room Name null — every row should produce a Room Name error
    result = _run_pipeline(sample_xlsx)
    errs = [i for i in result["validationIssues"] if i["severity"] == "error"]
    assert any(i["field"] == "Room Name" for i in errs)


def test_export_dynamic_writes_dynamic_columns(sample_xlsx: Path, tmp_path: Path) -> None:
    result = _run_pipeline(sample_xlsx)
    out = tmp_path / "out.xlsx"
    export_workbook(result, output_path=out, mode="dynamic_export")
    wb = load_workbook(out)
    assert "Hotel" in wb.sheetnames
    assert "Extraction Notes" in wb.sheetnames

    hotel_ws = wb["Hotel"]
    headers = [c.value for c in hotel_ws[1]]
    assert "Hotel Name" in headers
    assert "Country Code " in headers  # trailing space header preserved
    assert "SUPP-AI-CHILD" in headers
    # Dynamic child column should appear
    assert any(h and h.startswith("CHD(") for h in headers)
    # Internal columns must not appear in non-review modes
    assert "_source_refs" not in headers


def test_export_strict_writes_only_template_child_columns(sample_xlsx: Path, tmp_path: Path) -> None:
    result = _run_pipeline(sample_xlsx, mode="strict_template")
    out = tmp_path / "strict.xlsx"
    export_workbook(result, output_path=out, mode="strict_template")
    wb = load_workbook(out)
    hotel_ws = wb["Hotel"]
    headers = [c.value for c in hotel_ws[1]]

    # The strict template's fixed child columns should appear
    assert "CHD(0-2)" in headers
    assert "CHD(3-11)" in headers

    # Contract-specific child columns must NOT appear
    assert not any(h == "CHD(2-11.99)" for h in headers)


def test_export_review_includes_internal_columns(sample_xlsx: Path, tmp_path: Path) -> None:
    result = _run_pipeline(sample_xlsx)
    out = tmp_path / "review.xlsx"
    export_workbook(result, output_path=out, mode="dynamic_review", include_internal=True)
    wb = load_workbook(out)
    headers = [c.value for c in wb["Hotel"][1]]
    assert "_source_refs" in headers
    assert "_confidence" in headers
    assert "_warnings" in headers


def test_normalizer_moves_prose_in_numeric_field(single_hotel_xlsx: Path) -> None:
    parsed = parse_excel(single_hotel_xlsx)
    ir = build_ir_from_excel(parsed)
    options = {
        "supplierDefault": "X",
        "countryDefault": "EG",
        "currencyDefault": "EUR",
        "childColumnMode": "dynamic_review",
        "preserveChildPositions": True,
        "extractionMode": "text_only",
    }
    raw = stub_extract(ir, options)
    # Force a prose value into Charge
    raw["hotelRows"][0]["Charge"] = "tax included"
    normalized = normalize_result(raw, options, parsed["source_file"])
    assert normalized["hotelRows"][0]["Charge"] is None
    moved = [
        n
        for n in normalized["extractionNotes"]
        if "Charge" in n["Note"] and "tax included" in n["Note"]
    ]
    assert moved, "Prose should be moved to Extraction Notes"
