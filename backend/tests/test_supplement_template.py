"""Tests for the supplement-template writer.

Template format follows ``Hotel-Supplement-Import-Sample-AI.xlsx``
(sheet name 'Supplement', clean headers without typos).
"""
from __future__ import annotations

import openpyxl

from app.services.supplement_template import (
    SUPPLEMENT_HEADERS,
    write_supplement_rows,
    _template_path,
)


def test_template_has_bundled_columns_with_clean_headers() -> None:
    """Regression guard: openpyxl read-only mode TRUNCATES max_column on
    header-only templates, hiding leading columns. The writer must use the
    default load mode to see them. The bundled template predates the
    Min/Max Adult and Max Child columns (those are auto-appended at
    write time)."""
    wb = openpyxl.load_workbook(_template_path())
    assert "Supplement" in wb.sheetnames
    ws = wb["Supplement"]
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    assert "Supplement Code" in headers
    assert "Supplement Name" in headers
    assert "Supplement Type" in headers
    assert "Display on Customer Documentation" in headers
    assert "Standard / Count / Index" in headers
    # Old typo'd names must NOT appear.
    assert "Suppliment Code" not in headers
    assert "Suppliment Name" not in headers
    assert "Standard/Count/Index" not in headers


def test_write_empty_rows_produces_header_only_file(tmp_path) -> None:
    """No supplements -> header-only file (intentional, distinguishes
    'no supplements' from 'feature failed'). Output must carry the
    full SUPPLEMENT_HEADERS contract — including Min/Max Adult and
    Max Child auto-appended at write time."""
    out = tmp_path / "empty.xlsx"
    write_supplement_rows([], out)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb["Supplement"]
    assert ws.max_row == 1
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    for h in SUPPLEMENT_HEADERS:
        assert h in headers, f"header {h!r} missing from written supplement file"


def test_write_rows_round_trip(tmp_path) -> None:
    out = tmp_path / "rows.xlsx"
    write_supplement_rows(
        [
            {
                "Hotel Name": "Test Hotel",
                "Hotel Code": "TEST01",
                "Supplement Code": "HB-UP",
                "Supplement Name": "Half Board Upgrade",
                "Display on Customer Documentation": "Yes",
                "Display on Supplier Notification": "Yes",
                "Charge Type": "Per Person Per Night",
                "Calculation Method": "Standard",
                "Traveler Type": "Adult",
                "FareType Name": "Per Adult",
                "Standard / Count / Index": None,
                "Currency": "EUR",
                "Supplier Cost": 90,
                "Customer Price": 90,
                "Start Date (DD-MM-YYYY)": "01-05-2025",
                "End Date (DD-MM-YYYY)": "30-09-2025",
            }
        ],
        out,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["Supplement"]
    assert ws.max_row == 2
    row_by_header = {
        ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value
        for c in range(1, ws.max_column + 1)
    }
    assert row_by_header["Hotel Name"] == "Test Hotel"
    assert row_by_header["Supplement Name"] == "Half Board Upgrade"
    assert row_by_header["Display on Customer Documentation"] == "Yes"
    assert row_by_header["Standard / Count / Index"] is None
    assert row_by_header["Start Date (DD-MM-YYYY)"] == "01-05-2025"
    assert row_by_header["Supplier Cost"] == 90


def test_output_strips_inherited_header_colours(tmp_path) -> None:
    """The bundled template ships grey-fill + bold red headers — the
    writer must strip them so the output is plain Excel."""
    out = tmp_path / "plain.xlsx"
    write_supplement_rows(
        [{"Hotel Name": "X", "Charge Type": "Per Person Per Night",
          "Calculation Method": "Standard", "Traveler Type": "Adult"}],
        out,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["Supplement"]
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.fill.fill_type is None
    assert header_cell.font.color.rgb == "FF000000"
    body_cell = ws.cell(row=2, column=1)
    assert body_cell.fill.fill_type is None
    assert body_cell.font.color.rgb == "FF000000"


def test_unknown_headers_are_silently_dropped(tmp_path) -> None:
    out = tmp_path / "unknown.xlsx"
    write_supplement_rows(
        [{"Hotel Name": "X", "Bogus Column": "should not appear"}], out
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["Supplement"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Bogus Column" not in headers
    assert ws.cell(row=2, column=1).value == "X"
